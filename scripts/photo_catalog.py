"""
photo_catalog.py — Core engine for the PhotoCatalog application.

This module handles all the heavy lifting:
  - Scanning folders for supported image files
  - Extracting EXIF metadata via Pillow (PIL)
  - Extracting XMP metadata by parsing embedded XML
  - Converting GPS coordinates to decimal degrees
  - Parsing EXIF date strings into Python datetime objects
  - Building formatted Excel workbooks with Catalog and Summary sheets

Key Functions:
  scan_folder(folder)            — Recursively find all supported image files
  extract_metadata(filepath)     — Extract all EXIF + XMP data from one image
  write_excel(rows, path, name)  — Write metadata rows to a formatted .xlsx file

Dependencies:
  - Pillow (PIL) for image opening and EXIF tag reading
  - openpyxl for Excel workbook creation and formatting
  - xml.etree.ElementTree for XMP/XML parsing
"""
import os
import re
import sys
import struct
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
from collections import Counter

from PIL import Image
from PIL.ExifTags import TAGS, GPSTAGS, IFD
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# EXIF tag value lookups (duplicated from config.py for standalone use)
# ---------------------------------------------------------------------------
SUPPORTED_EXTENSIONS = {'.jpg', '.jpeg', '.tif', '.tiff', '.png', '.heif', '.heic', '.webp',
                        '.cr2', '.cr3', '.nef', '.arw', '.dng', '.orf', '.rw2'}

EXIF_FLASH_MODES = {0: 'No Flash', 1: 'Fired', 5: 'Fired, Return not detected',
                    7: 'Fired, Return detected', 8: 'On, Did not fire',
                    9: 'On, Fired', 16: 'Off, Did not fire', 24: 'Auto, Did not fire',
                    25: 'Auto, Fired', 32: 'No flash function'}
EXIF_ORIENTATION = {1: 'Horizontal', 2: 'Mirror horizontal', 3: 'Rotate 180',
                    4: 'Mirror vertical', 5: 'Mirror horizontal and rotate 270 CW',
                    6: 'Rotate 90 CW', 7: 'Mirror horizontal and rotate 90 CW', 8: 'Rotate 270 CW'}
EXIF_SCENE_CAPTURE = {0: 'Standard', 1: 'Landscape', 2: 'Portrait', 3: 'Night scene'}
EXIF_WHITE_BALANCE = {0: 'Auto', 1: 'Manual'}
EXIF_SENSING_METHOD = {1: 'Not defined', 2: 'One-chip color area', 3: 'Two-chip color area',
                       4: 'Three-chip color area', 5: 'Color sequential area',
                       7: 'Trilinear', 8: 'Color sequential linear'}
EXIF_COLOR_SPACE = {1: 'sRGB', 65535: 'Uncalibrated'}

# Claude changes 2026 03 14
# ---------------------------------------------------------------------------
# Date handling — Phase 2 enhancement
# ---------------------------------------------------------------------------
# These constants and functions convert EXIF date strings
# (e.g., "2024:01:15 14:30:00") into Python datetime objects so they
# can be stored as proper Excel dates (sortable, filterable) rather
# than plain text strings.
DATE_COLUMNS = {'DateTimeOriginal', 'File_Date'}

# EXIF date formats to try when parsing
EXIF_DATE_FORMATS = [
    '%Y:%m:%d %H:%M:%S',    # Standard EXIF: 2024:03:14 10:30:00
    '%Y-%m-%d %H:%M:%S',    # ISO-style:     2024-03-14 10:30:00
    '%Y:%m:%d',              # Date only:     2024:03:14
    '%Y-%m-%d',              # ISO date only: 2024-03-14
]


def parse_exif_date(value):
    """Try to parse an EXIF date string into a Python datetime object.
    Returns the datetime if successful, or the original string if not."""
    if not value or not isinstance(value, str):
        return value
    date_str = str(value).strip()[:19]
    for fmt in EXIF_DATE_FORMATS:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return value

# Windows + hidden folders that we never want to recurse into.
# Drive-root scans (C:\, J:\) typically hit $RECYCLE.BIN and System Volume
# Information which throw PermissionError if we try to read them.
SKIP_DIR_NAMES = {
    "$RECYCLE.BIN", "System Volume Information",
    ".git", "__pycache__", ".venv", "venv", "node_modules",
}


def _should_skip_dir(name):
    """Return True if this directory name should be pruned from the walk."""
    if name in SKIP_DIR_NAMES:
        return True
    # Hidden folders on both Unix (.name) and Mac/Win convention
    if name.startswith("."):
        return True
    return False


def scan_folder(folder_path):
    """
    Recursively scan a folder tree for supported image files.

    Walks the directory tree with ``os.walk`` and collects every file
    whose extension matches ``SUPPORTED_EXTENSIONS``. Skips Windows
    system folders (``$RECYCLE.BIN``, ``System Volume Information``),
    hidden directories (``.git``, ``.venv``, etc.), and tolerates
    permission errors by logging and continuing.

    Args:
        folder_path (str): Path to the root folder to scan.

    Returns:
        list[str]: Sorted list of absolute file paths for supported images.
    """
    files = []
    for root, dirs, names in os.walk(folder_path, onerror=None):
        # Prune skip-list directories in-place so os.walk doesn't descend.
        dirs[:] = [d for d in dirs if not _should_skip_dir(d)]
        for name in names:
            ext = Path(name).suffix.lower()
            if ext in SUPPORTED_EXTENSIONS:
                files.append(os.path.join(root, name))
    files.sort()
    return files


def prescan_folder(folder_path, progress_callback=None, cancel_event=None):
    """
    Walk a folder tree once and tally counts by file extension.

    Unlike :func:`scan_folder` this does not return the file paths —
    it's an informational pass used by the GUI's "Pre-Scan Folder"
    button to show users how many files, folders, and image types they
    have before committing to a full cataloging run. Only filesystem
    metadata is read, so it's ~100–1000× faster than a real catalog run.

    Args:
        folder_path (str): Path to the root folder to scan.
        progress_callback (callable, optional): Called periodically with
            ``(files_seen, folders_seen)`` so the GUI can update a
            running counter. Invoked every 250 files to keep signal
            traffic light.
        cancel_event (threading.Event, optional): If set partway through
            the walk, returns early with whatever has been counted so far.

    Returns:
        dict with keys:
            ``total_files``:     int — every file visited
            ``total_folders``:   int — every folder visited (excluding skipped)
            ``supported_count``: int — files matching SUPPORTED_EXTENSIONS
            ``other_count``:     int — everything else
            ``supported_by_ext``: Counter[str, int] — e.g. {'.jpg': 6892, '.heic': 1112}
            ``other_by_ext``:    Counter[str, int] — e.g. {'.mp4': 1230, '.pdf': 412}
            ``cancelled``:       bool — True if the scan was interrupted
    """
    supported_counter = Counter()
    other_counter = Counter()
    total_files = 0
    total_folders = 0
    cancelled = False

    for root, dirs, names in os.walk(folder_path, onerror=None):
        dirs[:] = [d for d in dirs if not _should_skip_dir(d)]
        total_folders += 1
        for name in names:
            total_files += 1
            ext = Path(name).suffix.lower()
            if ext in SUPPORTED_EXTENSIONS:
                supported_counter[ext] += 1
            elif ext:
                other_counter[ext] += 1
            else:
                # Files with no extension (README, Makefile, etc.)
                other_counter["(none)"] += 1

            # Progress throttled to every 250 files so we don't spam
            # Qt signals — still smooth enough at typical walk rates.
            if progress_callback is not None and total_files % 250 == 0:
                progress_callback(total_files, total_folders)

        if cancel_event is not None and cancel_event.is_set():
            cancelled = True
            break

    # One last callback so the final count shows even if we aren't on
    # the 250-file boundary.
    if progress_callback is not None:
        progress_callback(total_files, total_folders)

    return {
        "total_files": total_files,
        "total_folders": total_folders,
        "supported_count": sum(supported_counter.values()),
        "other_count": sum(other_counter.values()),
        "supported_by_ext": supported_counter,
        "other_by_ext": other_counter,
        "cancelled": cancelled,
    }


def convert_gps_to_decimal(gps_coords, gps_ref):
    if not gps_coords or not gps_ref:
        return None
    try:
        degrees = float(gps_coords[0])
        minutes = float(gps_coords[1])
        seconds = float(gps_coords[2])
        decimal = degrees + minutes / 60 + seconds / 3600
        if gps_ref in ('S', 'W'):
            decimal = -decimal
        return round(decimal, 6)
    except (TypeError, ValueError, IndexError):
        return None


# openpyxl rejects any string that contains ASCII control characters
# in these ranges and raises IllegalCharacterError on wb.save(). This
# shows up in the wild on EXIF fields from older cameras (e.g. Canon
# PowerShot SD1000 stores 'Canon PowerShot SD1000' followed by NUL
# padding in the Model tag), and it aborts the whole workbook write.
# The regex mirrors openpyxl's own ILLEGAL_CHARACTERS_RE.
_EXCEL_ILLEGAL_CHARS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')


# ---------------------------------------------------------------------------
# Concern severity markers — CR #1 ("FileDate fallback + File_Concern")
# ---------------------------------------------------------------------------
# Concerns are short messages, prefixed with [INFO] / [WARN] / [ERROR],
# that we surface into the File_Concern column and yellow/red-highlight
# at write time. Promoting previously-silent issues (sanitized EXIF,
# division-by-zero skips, stat failures, missing dates) into first-class
# data makes the workbook auditable and drives the rename-engine's
# validity gate.
CONCERN_INFO = "[INFO]"
CONCERN_WARN = "[WARN]"
CONCERN_ERROR = "[ERROR]"
CONCERN_MARKERS = (CONCERN_INFO, CONCERN_WARN, CONCERN_ERROR)

# Cell fill colors used by write_excel() when any of the above markers
# appear. INFO/WARN share one shade; ERROR gets a distinct shade so the
# eye can pick them out quickly when scrolling.
CONCERN_FILL_YELLOW = "FFFFE699"   # pale yellow for [INFO] / [WARN]
CONCERN_FILL_RED    = "FFFFC7CE"   # pale red for [ERROR]

# v3 — duplicate group highlight. Applied to the File_Name cell (not the
# whole row) whenever File_DupeGroup is populated, so visual scanning
# instantly surfaces duplicate clusters while leaving the rest of the
# row readable in its normal alternating fill.
DUPE_FILL_ORANGE    = "FFFFD59B"   # pale orange for duplicate group members


def _sanitize_cell_value(val, concerns=None, field=None):
    """
    Scrub control characters from strings before handing them to
    openpyxl. Leaves non-string types (datetime, int, float, None, etc.)
    untouched so Excel-native types still round-trip correctly.

    If *concerns* (a list) and *field* are supplied and sanitization
    actually changed the value, appends a ``[WARN]`` message describing
    which field was scrubbed.
    """
    if isinstance(val, str):
        # Strip NUL + other illegal control chars, then trim any trailing
        # whitespace those nulls were padding out.
        cleaned = _EXCEL_ILLEGAL_CHARS_RE.sub('', val).rstrip()
        if concerns is not None and field and cleaned != val:
            concerns.append(
                f"{CONCERN_WARN} Sanitized illegal characters in {field}"
            )
        return cleaned
    return val


def _safe_float(val, concerns=None, field=None):
    """
    Convert an EXIF value to float, returning None for anything that
    can't be coerced safely.

    Pillow represents rational EXIF fields (ExposureTime, FNumber, etc.)
    as ``IFDRational`` objects whose ``__float__`` raises
    ``ZeroDivisionError`` when the rational's denominator is 0 — a
    common malformation in scanner-produced JPEGs (passports, receipts,
    flatbed scans) where EXIF is stripped or half-populated. Catching it
    here lets the catalog keep running and simply leave the field blank
    for those files rather than failing the whole run.

    If *concerns* (a list) and *field* are supplied and we hit a
    ``ZeroDivisionError``, appends a ``[WARN]`` message so the row is
    flagged in File_Concern instead of silently dropping the value.
    """
    if val is None:
        return None
    try:
        return float(val)
    except ZeroDivisionError:
        if concerns is not None and field:
            concerns.append(
                f"{CONCERN_WARN} Skipped EXIF {field} \u2014 division by zero"
            )
        return None
    except (TypeError, ValueError):
        return None


def format_exposure_time(val):
    """
    Format an exposure time value as a human-readable string.

    EXIF stores exposure as a float (e.g., 0.004). This converts it to
    a fraction like "1/250" for short exposures, or a decimal like "2.0"
    for long exposures.

    Args:
        val: The raw exposure time value from EXIF data.

    Returns:
        str: Formatted exposure string, or the original value as string.
    """
    if val is None:
        return None
    try:
        val = float(val)
        if val == 0:
            # Some scanner/phone JPEGs store ExposureTime == 0. Treat
            # it as "not available" rather than dividing by zero below.
            return None
        if val < 1:
            denom = round(1 / val)
            return f"1/{denom}"
        return f"{val:.1f}"
    except (TypeError, ValueError, ZeroDivisionError):
        return str(val)


def format_lens_spec(spec):
    """
    Format a lens specification tuple into a readable string.

    EXIF stores lens info as a 4-element tuple:
    (min_focal, max_focal, min_aperture, max_aperture).
    This formats it as "18.0-55.0mm f/3.5-5.6".

    Args:
        spec: Tuple of (min_focal, max_focal, min_fstop, max_fstop).

    Returns:
        str or None: Formatted lens spec, or None if input is empty.
    """
    if not spec:
        return None
    try:
        return f"{float(spec[0]):.1f}-{float(spec[1]):.1f}mm f/{float(spec[2]):.1f}-{float(spec[3]):.1f}"
    except (TypeError, ValueError, IndexError, ZeroDivisionError):
        return str(spec)


def extract_exif(filepath, concerns=None):
    """
    Extract EXIF metadata from an image file using Pillow.

    Opens the image, reads its EXIF data, and maps raw tag values to
    human-readable column names. Handles special processing for:
      - GPS coordinates (converted to decimal degrees)
      - Exposure time (formatted as fractions)
      - Lookup values (flash mode, metering, orientation, etc.)

    Args:
        filepath (str): Path to the image file.
        concerns (list, optional): If supplied, receives ``[WARN]`` /
            ``[ERROR]`` strings describing any EXIF values we had to
            skip (e.g. division-by-zero rationals). Caller is expected
            to surface these via the File_Concern column.

    Returns:
        dict: Key-value pairs of extracted metadata fields.
              Keys are column names (e.g., 'CameraMake', 'ISO').
    """
    data = {}
    try:
        img = Image.open(filepath)
    except Exception:
        return data

    data['ImageWidth'] = img.width
    data['ImageHeight'] = img.height

    try:
        exif_data = img._getexif()
    except Exception:
        exif_data = None

    if not exif_data:
        return data

    gps_info = {}
    for tag_id, value in exif_data.items():
        tag = TAGS.get(tag_id, str(tag_id))

        if tag == 'GPSInfo':
            for gps_tag_id, gps_val in value.items():
                gps_tag = GPSTAGS.get(gps_tag_id, str(gps_tag_id))
                gps_info[gps_tag] = gps_val
            continue

        if tag == 'MakerNote':
            continue

        if isinstance(value, bytes):
            try:
                value = value.decode('utf-8', errors='ignore').strip('\x00')
            except Exception:
                continue

        # Map and format known fields
        if tag == 'Make':
            data['CameraMake'] = str(value).strip()
        elif tag == 'Model':
            data['CameraModel'] = str(value).strip()
        elif tag == 'Software':
            data['Software'] = str(value).strip()
        elif tag == 'LensMake':
            data['LensMake'] = str(value).strip()
        elif tag == 'LensModel':
            data['LensModel'] = str(value).strip()
        elif tag == 'LensSpecification':
            data['LensSpec'] = format_lens_spec(value)
        elif tag == 'DateTimeOriginal':
            data['DateTimeOriginal'] = str(value)
        elif tag == 'OffsetTime':
            data['OffsetTime'] = str(value)
        elif tag == 'OffsetTimeOriginal':
            data['OffsetTimeOriginal'] = str(value)
        elif tag == 'OffsetTimeDigitized':
            data['OffsetTimeDigitized'] = str(value)
        elif tag == 'ExposureTime':
            data['ExposureTime'] = format_exposure_time(value)
        elif tag == 'FNumber':
            fv = _safe_float(value, concerns, 'FNumber')
            data['FNumber'] = round(fv, 2) if fv is not None else None
        elif tag == 'ISOSpeedRatings':
            fv = _safe_float(value, concerns, 'ISO')
            data['ISO'] = int(fv) if fv is not None else None
        elif tag == 'ShutterSpeedValue':
            fv = _safe_float(value, concerns, 'ShutterSpeed')
            data['ShutterSpeed'] = round(fv, 2) if fv is not None else None
        elif tag == 'ApertureValue':
            fv = _safe_float(value, concerns, 'Aperture')
            data['Aperture'] = round(fv, 2) if fv is not None else None
        elif tag == 'BrightnessValue':
            fv = _safe_float(value, concerns, 'Brightness')
            data['Brightness'] = round(fv, 2) if fv is not None else None
        elif tag == 'ExposureBiasValue':
            fv = _safe_float(value, concerns, 'ExposureBias')
            data['ExposureBias'] = round(fv, 2) if fv is not None else None
        elif tag == 'Flash':
            data['Flash'] = EXIF_FLASH_MODES.get(value, str(value))
        elif tag == 'FocalLength':
            fv = _safe_float(value, concerns, 'FocalLength_mm')
            data['FocalLength_mm'] = round(fv, 2) if fv is not None else None
        elif tag == 'FocalLengthIn35mmFilm':
            fv = _safe_float(value, concerns, 'FocalLength35mm')
            data['FocalLength35mm'] = int(fv) if fv is not None else None
        elif tag == 'WhiteBalance':
            data['WhiteBalance'] = EXIF_WHITE_BALANCE.get(value, str(value))
        elif tag == 'SceneCaptureType':
            data['SceneCaptureType'] = EXIF_SCENE_CAPTURE.get(value, str(value))
        elif tag == 'SensingMethod':
            data['SensingMethod'] = EXIF_SENSING_METHOD.get(value, str(value))
        elif tag == 'ColorSpace':
            data['ColorSpace'] = EXIF_COLOR_SPACE.get(value, str(value))
        elif tag == 'Orientation':
            data['Orientation'] = EXIF_ORIENTATION.get(value, str(value))
        elif tag == 'CompositeImage':
            data['CompositeImage'] = int(value) if value else None
        elif tag in ('ExifImageWidth', 'ExifImageHeight'):
            pass  # Already captured from img.width/height
        elif tag in ('ExifOffset', 'ExifVersion', 'ComponentsConfiguration',
                      'FlashPixVersion', 'YCbCrPositioning', 'SceneType',
                      'ResolutionUnit', 'HostComputer'):
            if tag == 'HostComputer':
                data['HostComputer'] = str(value).strip()
        elif tag == 'XResolution':
            fv = _safe_float(value, concerns, 'XResolution')
            data['XResolution'] = int(fv) if fv is not None else None
        elif tag == 'YResolution':
            fv = _safe_float(value, concerns, 'YResolution')
            data['YResolution'] = int(fv) if fv is not None else None
        elif tag == 'SubjectLocation':
            data['SubjectLocation'] = str(value)

    # Process GPS
    if gps_info:
        lat = convert_gps_to_decimal(gps_info.get('GPSLatitude'), gps_info.get('GPSLatitudeRef'))
        lon = convert_gps_to_decimal(gps_info.get('GPSLongitude'), gps_info.get('GPSLongitudeRef'))
        if lat is not None:
            data['GPSLatitude'] = lat
        if lon is not None:
            data['GPSLongitude'] = lon
        if 'GPSAltitude' in gps_info:
            try:
                alt = float(gps_info['GPSAltitude'])
                ref = gps_info.get('GPSAltitudeRef', 0)
                if ref == 1:
                    alt = -alt
                data['GPSAltitude_m'] = round(alt, 1)
            except (TypeError, ValueError, ZeroDivisionError):
                pass
        if 'GPSSpeed' in gps_info:
            try:
                data['GPSSpeed'] = round(float(gps_info['GPSSpeed']), 2)
            except (TypeError, ValueError, ZeroDivisionError):
                pass
        if 'GPSImgDirection' in gps_info:
            try:
                data['GPSImgDirection'] = round(float(gps_info['GPSImgDirection']), 2)
            except (TypeError, ValueError, ZeroDivisionError):
                pass
        if 'GPSDateStamp' in gps_info:
            data['GPSDateStamp'] = str(gps_info['GPSDateStamp'])

    return data


def extract_xmp(filepath):
    """
    Extract raw XMP/XML metadata string from an image file.

    XMP data is embedded as an XML document inside the image file,
    delimited by '<x:xmpmeta' and '</x:xmpmeta>' markers. This function
    reads the file in binary mode and searches for these markers.

    Args:
        filepath (str): Path to the image file.

    Returns:
        str or None: The XMP XML string if found, None otherwise.
    """
    data = {}
    try:
        with open(filepath, 'rb') as f:
            raw = f.read()
    except Exception:
        return data

    xmp_start = raw.find(b'<x:xmpmeta')
    xmp_end = raw.find(b'</x:xmpmeta>')
    if xmp_start < 0 or xmp_end < 0:
        return data

    xmp_bytes = raw[xmp_start:xmp_end + len(b'</x:xmpmeta>')]
    try:
        xmp_str = xmp_bytes.decode('utf-8', errors='ignore')
    except Exception:
        return data

    ns = {
        'x': 'adobe:ns:meta/',
        'rdf': 'http://www.w3.org/1999/02/22-rdf-syntax-ns#',
        'xmp': 'http://ns.adobe.com/xap/1.0/',
        'exif': 'http://ns.adobe.com/exif/1.0/',
        'photoshop': 'http://ns.adobe.com/photoshop/1.0/',
        'dc': 'http://purl.org/dc/elements/1.1/',
        'mwg-rs': 'http://www.metadataworkinggroup.com/schemas/regions/',
        'stArea': 'http://ns.adobe.com/xmp/sType/Area#',
        'stDim': 'http://ns.adobe.com/xap/1.0/sType/Dimensions#',
        'apple-fi': 'http://ns.apple.com/faceinfo/1.0/',
        'lr': 'http://ns.adobe.com/lightroom/1.0/',
        'tiff': 'http://ns.adobe.com/tiff/1.0/',
        'crs': 'http://ns.adobe.com/camera-raw-settings/1.0/',
    }

    try:
        root = ET.fromstring(xmp_str)
    except ET.ParseError:
        return data

    # Find rdf:Description elements
    for desc in root.iter('{http://www.w3.org/1999/02/22-rdf-syntax-ns#}Description'):
        # Read attributes
        for attr_name, attr_val in desc.attrib.items():
            clean = attr_name.split('}')[-1] if '}' in attr_name else attr_name
            ns_prefix = ''
            if '}' in attr_name:
                ns_uri = attr_name.split('}')[0].lstrip('{')
                for prefix, uri in ns.items():
                    if uri == ns_uri:
                        ns_prefix = prefix + ':'
                        break
            if clean in ('about', 'parseType'):
                continue
            col_name = f"XMP_{ns_prefix}{clean}"
            data[col_name] = attr_val

    # Count face regions
    face_count = 0
    for region in root.iter('{http://www.metadataworkinggroup.com/schemas/regions/}Type'):
        face_count += 1
    # Better: count Description elements with mwg-rs:Type="Face"
    face_count = 0
    for desc_el in root.iter('{http://www.w3.org/1999/02/22-rdf-syntax-ns#}Description'):
        type_attr = desc_el.get('{http://www.metadataworkinggroup.com/schemas/regions/}Type')
        if type_attr == 'Face':
            face_count += 1
    if face_count > 0:
        data['XMP_FaceRegionCount'] = face_count

    # Check for regions element
    regions = root.find('.//{http://www.metadataworkinggroup.com/schemas/regions/}Regions')
    if regions is not None:
        data['XMP_HasFaceRegions'] = 'Yes'

    return data


def extract_metadata(filepath):
    """
    Build one catalog row from a single image file.

    The row uses the ``File_`` prefix on every app-generated / filesystem-
    sourced column (see CR #1) so they group visually to the left of the
    camera-sourced EXIF/XMP fields.

    The row carries an internal ``_concerns`` list (not a column) that
    write_excel() flattens into the ``File_Concern`` cell — severity
    markers ``[INFO]`` / ``[WARN]`` / ``[ERROR]`` drive row highlighting
    and, in a later phase, rename/move gating.
    """
    filename = os.path.basename(filepath)
    file_size = os.path.getsize(filepath)
    if file_size > 1024 * 1024:
        size_str = f"{file_size / (1024 * 1024):.1f} MB"
    else:
        size_str = f"{file_size / 1024:.0f} KB"

    # Normalize extension to lowercase-with-dot (e.g. '.jpg', '.heic') so
    # the column sorts/filters cleanly regardless of on-disk case
    # ('IMG_6596.JPG' and 'img_6596.jpg' both become '.jpg').
    file_ext = Path(filename).suffix.lower()

    concerns: list = []

    # File_Date — capture the filesystem Modified timestamp so the
    # rename engine can fall back on it when DateTimeOriginal is blank
    # (and, more generally, so users who filter by "real shoot date"
    # have something to fall back on for scanned/processed files).
    file_date = None
    try:
        file_date = datetime.fromtimestamp(os.stat(filepath).st_mtime)
    except OSError as e:
        concerns.append(
            f"{CONCERN_ERROR} Could not read file Modified date: {e}"
        )

    row = {
        'File_Name': filename,
        'File_Extension': file_ext,
        # File_RenameName is populated later by the rename engine — kept
        # in the row dict (and in COLUMN_ORDER) as a blank placeholder
        # so the column always exists in the workbook even before the
        # user runs a rename pass.
        'File_RenameName': None,
        'File_Path': filepath,
        'File_Size': size_str,
        'File_SizeBytes': file_size,
        'File_Date': file_date,
        # v3 placeholders — populated by later pipeline passes. Keeping
        # them in the row dict up front ensures the columns always exist
        # in the workbook even for pipelines that skip hashing / dupe
        # detection / destination composition.
        'File_Hash': None,          # MD5, filled by duplicate_detector.populate_hashes (hash mode only)
        'File_DupeGroup': None,     # int group id, filled by duplicate_detector.detect_duplicates
        'File_DupeKeep': None,      # bool keeper flag (True/False/None), same
        'File_DestFolder': None,    # relative dest subfolder, filled by copy_engine.populate_destination_columns
        'File_DestPath': None,      # full dest path, same
        'File_Status': None,        # pending/copied/skipped/dupe_moved/dupe_deleted, same
        # File_Concern is written as the final join of _concerns at
        # workbook-write time; placeholder here keeps the column order.
        'File_Concern': None,
        '_concerns': concerns,
    }

    exif = extract_exif(filepath, concerns=concerns)
    row.update(exif)

    xmp = extract_xmp(filepath)
    row.update(xmp)

    return row


# --- Preferred column order ---
# App-generated / filesystem-sourced columns use the File_ prefix and
# group together on the left of the workbook so they visually separate
# from EXIF/XMP camera-sourced fields. File_Concern sits just before
# File_Path so it reads as a "flags" column paired with the source path.
COLUMN_ORDER = [
    'File_Name', 'File_Extension', 'File_RenameName',
    'File_Size', 'File_SizeBytes', 'File_Date',
    # --- v3 additions: hash, duplicate grouping, destination mapping,
    # and per-row operation status. Slotted between File_Date and
    # File_Concern so the expanded File_ block still groups at the
    # left edge of the sheet.
    'File_Hash',
    'File_DupeGroup', 'File_DupeKeep',
    'File_DestFolder', 'File_DestPath', 'File_Status',
    'File_Concern', 'File_Path',
    'ImageWidth', 'ImageHeight',
    'CameraMake', 'CameraModel', 'HostComputer', 'Software',
    'LensMake', 'LensModel', 'LensSpec',
    'DateTimeOriginal',
    'OffsetTime', 'OffsetTimeOriginal', 'OffsetTimeDigitized',
    'ExposureTime', 'FNumber', 'ISO',
    'ShutterSpeed', 'Aperture', 'Brightness', 'ExposureBias',
    'Flash', 'FocalLength_mm', 'FocalLength35mm',
    'WhiteBalance', 'SceneCaptureType', 'SensingMethod', 'ColorSpace',
    'CompositeImage', 'Orientation',
    'XResolution', 'YResolution', 'SubjectLocation',
    'GPSLatitude', 'GPSLongitude', 'GPSAltitude_m',
    'GPSSpeed', 'GPSImgDirection', 'GPSDateStamp',
    'FaceCount_Detected', 'PersonNames',
]

# Row-dict keys that are carried through the pipeline but should never
# be written as a workbook column (they're internal plumbing).
_INTERNAL_ROW_KEYS = {'_concerns'}


def build_column_list(all_rows):
    all_keys = set()
    for row in all_rows:
        all_keys.update(row.keys())
    all_keys -= _INTERNAL_ROW_KEYS

    ordered = [c for c in COLUMN_ORDER if c in all_keys]
    remaining = sorted(all_keys - set(ordered))
    return ordered + remaining


def _tally_concern_severity(concerns):
    """
    Count occurrences of each severity marker across *concerns* (a list
    of strings). Returns a dict with keys 'info', 'warn', 'error'.
    """
    info = warn = error = 0
    for msg in concerns:
        if CONCERN_ERROR in msg:
            error += 1
        if CONCERN_WARN in msg:
            warn += 1
        if CONCERN_INFO in msg:
            info += 1
    return {'info': info, 'warn': warn, 'error': error}


def write_excel(all_rows, output_path, folder_name):
    """
    Write all metadata to a formatted Excel workbook.

    Creates a workbook with two sheets:
      - Catalog: One row per photo with all metadata columns, formatted with
        headers, alternating row colors, auto-filter, and frozen header row.
        Date columns use proper Excel date formatting (Phase 2 enhancement).
        The File_Concern column is populated from each row's internal
        ``_concerns`` list (sanitization warnings, division-by-zero skips,
        fallback-date notices, stat-failure errors) and the cell is filled
        with a pale yellow (info/warn) or pale red (error) shade based on
        the highest severity present.
      - Summary: Statistics overview including photo count, date range,
        camera/lens breakdown, exposure stats, GPS coverage, and face data.

    Args:
        all_rows (list[dict]): Metadata rows from extract_metadata().
        output_path (str): Full path for the output .xlsx file.
        folder_name (str): Name of the scanned folder (shown in Summary).

    Returns:
        tuple[int, int, dict]: ``(number_of_columns, number_of_rows, concern_totals)``
        where ``concern_totals`` has keys ``info``, ``warn``, ``error``,
        ``rows_with_concerns``, and ``total_rows``.
    """
    columns = build_column_list(all_rows)

    wb = Workbook()

    # --- CATALOG SHEET ---
    ws = wb.active
    ws.title = "Catalog"

    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', fgColor='2F5496')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_font = Font(name='Arial', size=10)
    cell_align = Alignment(vertical='top')
    thin_border = Border(
        bottom=Side(style='thin', color='D9E2F3')
    )
    alt_fill = PatternFill('solid', fgColor='F2F2F2')
    concern_fill_yellow = PatternFill('solid', fgColor=CONCERN_FILL_YELLOW)
    concern_fill_red = PatternFill('solid', fgColor=CONCERN_FILL_RED)
    dupe_fill_orange = PatternFill('solid', fgColor=DUPE_FILL_ORANGE)

    # Find the File_Concern column index (if present) so we can skip it
    # in the main write loop and populate it once per row after all
    # other cells are in place (sanitization during the loop appends
    # entries to row['_concerns'], and we want those included).
    try:
        concern_col_idx = columns.index('File_Concern') + 1
    except ValueError:
        concern_col_idx = None

    # v3 — locate the File_Name column so we can paint the pale-orange
    # dupe fill on exactly that cell for rows belonging to a duplicate
    # group. File_Name is always present in v3 but we still guard with
    # a ValueError fallback for paranoia's sake.
    try:
        file_name_col_idx = columns.index('File_Name') + 1
    except ValueError:
        file_name_col_idx = None

    # Write headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Claude Suggested Change 2026 03 14
    # Excel date format string
    date_number_format = 'YYYY-MM-DD HH:MM:SS'

    # Aggregate severity counter rolled up across the whole run so
    # run_catalog can log a one-line summary to the user.
    all_concerns: list = []
    rows_with_concerns = 0

    # Write data
    for row_idx, row_data in enumerate(all_rows, 2):
        concerns = row_data.get('_concerns') or []
        for col_idx, col_name in enumerate(columns, 1):
            # File_Concern is populated last (after sanitization may add
            # more entries). Skip here so the cell's final value reflects
            # the full list.
            if col_idx == concern_col_idx:
                # Still write the alt-fill background on the placeholder
                # cell so the column isn't visually blank mid-run — the
                # concern-fill below will override this if applicable.
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = cell_font
                cell.alignment = cell_align
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
                continue

            val = row_data.get(col_name)

            # Convert date columns to proper Excel datetime values.
            # File_Date is already a datetime from extract_metadata;
            # DateTimeOriginal arrives as a string and needs parsing.
            if col_name in DATE_COLUMNS and val is not None and not isinstance(val, datetime):
                val = parse_exif_date(val)

            # Scrub NUL / control chars from strings so openpyxl doesn't
            # raise IllegalCharacterError on quirky camera EXIF padding.
            # Pass in the row's concerns list so scrubbed fields get a
            # [WARN] entry and the row is flagged in File_Concern.
            val = _sanitize_cell_value(val, concerns=concerns, field=col_name)

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = cell_font

            # Apply date format if this is a date column with a datetime value
            if col_name in DATE_COLUMNS and isinstance(val, datetime):
                cell.number_format = date_number_format
            cell.alignment = cell_align
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

        # Flatten the row's concerns into the File_Concern cell and
        # pick a fill color based on the highest severity present.
        if concern_col_idx is not None:
            cell = ws.cell(row=row_idx, column=concern_col_idx)
            if concerns:
                rows_with_concerns += 1
                all_concerns.extend(concerns)
                cell.value = "; ".join(concerns)
                if any(CONCERN_ERROR in c for c in concerns):
                    cell.fill = concern_fill_red
                else:
                    # [INFO] and [WARN] share the same pale yellow shade.
                    cell.fill = concern_fill_yellow

        # v3 — paint the File_Name cell pale orange if this row is a
        # member of a duplicate group. Applied AFTER the alt-fill write
        # above so it reliably overrides the zebra striping for dupes.
        if file_name_col_idx is not None and row_data.get('File_DupeGroup'):
            ws.cell(row=row_idx, column=file_name_col_idx).fill = dupe_fill_orange

    # Auto-fit column widths
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(str(col_name))
        for row_idx in range(2, min(len(all_rows) + 2, 52)):  # Sample first 50 rows
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 50))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    # Freeze header row and enable auto-filter
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(all_rows) + 1}"

    # --- SUMMARY SHEET ---
    ws2 = wb.create_sheet("Summary")
    title_font = Font(name='Arial', bold=True, size=14, color='2F5496')
    label_font = Font(name='Arial', bold=True, size=11)
    value_font = Font(name='Arial', size=11)
    section_font = Font(name='Arial', bold=True, size=12, color='2F5496')
    section_fill = PatternFill('solid', fgColor='D9E2F3')

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 50

    r = 1
    ws2.cell(row=r, column=1, value="Photo Catalog Summary").font = title_font
    ws2.cell(row=r, column=1).alignment = Alignment(horizontal='left')
    r += 1
    ws2.cell(row=r, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = value_font
    r += 2

    # General stats
    ws2.cell(row=r, column=1, value="General").font = section_font
    ws2.cell(row=r, column=1).fill = section_fill
    ws2.cell(row=r, column=2).fill = section_fill
    r += 1

    ws2.cell(row=r, column=1, value="Folder").font = label_font
    ws2.cell(row=r, column=2, value=folder_name).font = value_font
    r += 1
    photo_count = len(all_rows)
    ws2.cell(row=r, column=1, value="Total Photos").font = label_font
    ws2.cell(row=r, column=2, value=photo_count).font = value_font
    r += 1

    # Date range — prefer DateTimeOriginal, fall back to File_Date so
    # rows missing EXIF still contribute to the span (CR #1).
    dates = []
    for row in all_rows:
        dt = row.get('DateTimeOriginal') or row.get('File_Date')
        if not dt:
            continue
        if isinstance(dt, datetime):
            dates.append(dt)
        else:
            try:
                dates.append(datetime.strptime(str(dt)[:19], '%Y:%m:%d %H:%M:%S'))
            except ValueError:
                pass
    if dates:
        ws2.cell(row=r, column=1, value="Earliest Photo").font = label_font
        ws2.cell(row=r, column=2, value=min(dates).strftime('%Y-%m-%d %H:%M:%S')).font = value_font
        r += 1
        ws2.cell(row=r, column=1, value="Latest Photo").font = label_font
        ws2.cell(row=r, column=2, value=max(dates).strftime('%Y-%m-%d %H:%M:%S')).font = value_font
        r += 1
        span = max(dates) - min(dates)
        ws2.cell(row=r, column=1, value="Date Span").font = label_font
        ws2.cell(row=r, column=2, value=f"{span.days} days").font = value_font
        r += 1
    r += 1

    # Camera stats
    ws2.cell(row=r, column=1, value="Cameras & Lenses").font = section_font
    ws2.cell(row=r, column=1).fill = section_fill
    ws2.cell(row=r, column=2).fill = section_fill
    r += 1

    cameras = Counter(row.get('CameraModel', 'Unknown') for row in all_rows if row.get('CameraModel'))
    for cam, count in cameras.most_common(10):
        ws2.cell(row=r, column=1, value=cam).font = value_font
        ws2.cell(row=r, column=2, value=f"{count} photos").font = value_font
        r += 1

    lenses = Counter(row.get('LensModel', 'Unknown') for row in all_rows if row.get('LensModel'))
    if lenses:
        for lens, count in lenses.most_common(10):
            ws2.cell(row=r, column=1, value=lens).font = value_font
            ws2.cell(row=r, column=2, value=f"{count} photos").font = value_font
            r += 1
    r += 1

    # Exposure stats
    ws2.cell(row=r, column=1, value="Exposure Stats").font = section_font
    ws2.cell(row=r, column=1).fill = section_fill
    ws2.cell(row=r, column=2).fill = section_fill
    r += 1

    isos = [row['ISO'] for row in all_rows if row.get('ISO')]
    if isos:
        ws2.cell(row=r, column=1, value="ISO Range").font = label_font
        ws2.cell(row=r, column=2, value=f"{min(isos)} – {max(isos)} (avg {sum(isos)//len(isos)})").font = value_font
        r += 1

    fnums = [row['FNumber'] for row in all_rows if row.get('FNumber')]
    if fnums:
        ws2.cell(row=r, column=1, value="Aperture Range").font = label_font
        ws2.cell(row=r, column=2, value=f"f/{min(fnums)} – f/{max(fnums)}").font = value_font
        r += 1

    focals = [row['FocalLength35mm'] for row in all_rows if row.get('FocalLength35mm')]
    if focals:
        ws2.cell(row=r, column=1, value="Focal Length Range (35mm eq.)").font = label_font
        ws2.cell(row=r, column=2, value=f"{min(focals)}mm – {max(focals)}mm").font = value_font
        r += 1
    r += 1

    # GPS stats
    ws2.cell(row=r, column=1, value="GPS Coverage").font = section_font
    ws2.cell(row=r, column=1).fill = section_fill
    ws2.cell(row=r, column=2).fill = section_fill
    r += 1
    gps_count = sum(1 for row in all_rows if row.get('GPSLatitude'))
    ws2.cell(row=r, column=1, value="Photos with GPS").font = label_font
    pct = f"{gps_count / photo_count * 100:.0f}%" if photo_count else "0%"
    ws2.cell(row=r, column=2, value=f"{gps_count} of {photo_count} ({pct})").font = value_font
    r += 1

    # Face detection (XMP metadata)
    xmp_face_count = sum(1 for row in all_rows if row.get('XMP_FaceRegionCount'))
    if xmp_face_count:
        r += 1
        ws2.cell(row=r, column=1, value="Face Detection (XMP Metadata)").font = section_font
        ws2.cell(row=r, column=1).fill = section_fill
        ws2.cell(row=r, column=2).fill = section_fill
        r += 1
        ws2.cell(row=r, column=1, value="Photos with XMP Faces").font = label_font
        ws2.cell(row=r, column=2, value=f"{xmp_face_count} photos").font = value_font
        r += 1
        total_xmp_faces = sum(row.get('XMP_FaceRegionCount', 0) for row in all_rows)
        ws2.cell(row=r, column=1, value="Total XMP Faces").font = label_font
        ws2.cell(row=r, column=2, value=total_xmp_faces).font = value_font
        r += 1

    # Face recognition (OpenCV detection + clustering)
    detected_face_count = sum(1 for row in all_rows if row.get('FaceCount_Detected', 0) > 0)
    if detected_face_count:
        r += 1
        ws2.cell(row=r, column=1, value="Face Recognition (OpenCV)").font = section_font
        ws2.cell(row=r, column=1).fill = section_fill
        ws2.cell(row=r, column=2).fill = section_fill
        r += 1
        ws2.cell(row=r, column=1, value="Photos with Detected Faces").font = label_font
        ws2.cell(row=r, column=2, value=f"{detected_face_count} photos").font = value_font
        r += 1
        total_detected = sum(row.get('FaceCount_Detected', 0) for row in all_rows)
        ws2.cell(row=r, column=1, value="Total Faces Detected").font = label_font
        ws2.cell(row=r, column=2, value=total_detected).font = value_font
        r += 1

        # Count unique persons
        all_persons = set()
        for row in all_rows:
            names = row.get('PersonNames', '')
            if names:
                for p in names.split(', '):
                    all_persons.add(p.strip())
        ws2.cell(row=r, column=1, value="Unique Persons Identified").font = label_font
        ws2.cell(row=r, column=2, value=len(all_persons)).font = value_font
        r += 1

        # Top persons by photo count
        person_counter = Counter()
        for row in all_rows:
            names = row.get('PersonNames', '')
            if names:
                for p in names.split(', '):
                    person_counter[p.strip()] += 1
        r += 1
        ws2.cell(row=r, column=1, value="Top Persons by Photo Count").font = label_font
        r += 1
        for person, count in person_counter.most_common(15):
            ws2.cell(row=r, column=1, value=f"  {person}").font = value_font
            ws2.cell(row=r, column=2, value=f"{count} photos").font = value_font
            r += 1

    wb.save(output_path)

    # Aggregate severity tally so run_catalog can log a summary line.
    concern_totals = _tally_concern_severity(all_concerns)
    concern_totals['rows_with_concerns'] = rows_with_concerns
    concern_totals['total_rows'] = photo_count
    return len(columns), photo_count, concern_totals


def main():
    # Parse arguments
    enable_faces = '--no-faces' not in sys.argv
    folder = None
    for arg in sys.argv[1:]:
        if not arg.startswith('--'):
            folder = arg
    if not folder:
        folder = '/sessions/relaxed-gallant-goodall/mnt/Claude Photos'

    folder_name = os.path.basename(folder.rstrip('/'))
    today = datetime.now().strftime('%Y-%m-%d')
    output_filename = f"PhotoCatalog_{folder_name}_{today}.xlsx"
    output_path = os.path.join(folder, output_filename)

    print(f"Scanning: {folder}")
    files = scan_folder(folder)
    print(f"Found {len(files)} supported image files")

    # Phase 1: Extract EXIF/XMP metadata
    print("\n--- Phase 1: Extracting metadata ---")
    all_rows = []
    for i, filepath in enumerate(files):
        if (i + 1) % 50 == 0 or i == 0:
            print(f"  Processing {i + 1}/{len(files)}: {os.path.basename(filepath)}")
        row = extract_metadata(filepath)
        all_rows.append(row)

    # Phase 2: Face recognition (optional)
    if enable_faces:
        print("\n--- Phase 2: Face recognition ---")
        try:
            from face_recognition import process_all_faces

            def progress(current, total, name):
                print(f"  Detecting faces {current}/{total}: {name}")

            face_results = process_all_faces(files, progress_callback=progress)

            # Merge face results into metadata rows
            for i, filepath in enumerate(files):
                result = face_results.get(filepath, {})
                all_rows[i]['FaceCount_Detected'] = result.get('face_count', 0)
                all_rows[i]['PersonNames'] = result.get('person_names', '')

            total_faces = sum(r.get('face_count', 0) for r in face_results.values())
            photos_with_faces = sum(1 for r in face_results.values() if r.get('face_count', 0) > 0)
            all_persons = set()
            for r in face_results.values():
                names = r.get('person_names', '')
                if names:
                    for p in names.split(', '):
                        all_persons.add(p.strip())
            print(f"  Found {total_faces} faces in {photos_with_faces} photos")
            print(f"  Identified {len(all_persons)} unique persons")
        except ImportError:
            print("  Warning: face_recognition module not found, skipping face detection")
        except Exception as e:
            print(f"  Warning: Face recognition failed: {e}")
    else:
        print("\n--- Phase 2: Face recognition skipped (--no-faces) ---")

    # Phase 3: Write Excel
    print(f"\n--- Phase 3: Writing Excel ---")
    print(f"Output: {output_path}")
    num_cols, num_rows, concern_totals = write_excel(all_rows, output_path, folder_name)
    print(f"Done! {num_rows} photos \u00d7 {num_cols} columns")
    if concern_totals['rows_with_concerns']:
        print(
            f"File_Concern markers: "
            f"[INFO] {concern_totals['info']:,}, "
            f"[WARN] {concern_totals['warn']:,}, "
            f"[ERROR] {concern_totals['error']:,} "
            f"across {concern_totals['rows_with_concerns']:,} row(s)"
        )


if __name__ == '__main__':
    main()
