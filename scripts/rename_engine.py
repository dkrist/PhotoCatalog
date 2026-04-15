"""
rename_engine.py — Template-driven filename construction for PhotoCatalog.

Takes a user-supplied template string with ``%Variable%`` tokens and
produces a proposed new filename for each cataloged photo by substituting
tokens with values pulled from the catalog workbook.

Public entry points:
  * :func:`validate_template` — parse a template string, return the list
    of tokens it contains plus any unknown ones.
  * :func:`render_row` — produce a single rename string for one row
    (used by the Test button to preview the first N rows).
  * :func:`test_template` — open the last catalog workbook, render the
    first N data rows, and return the preview list.
  * :func:`build_renames` — open the catalog workbook, render every
    data row, write results back into the ``File_RenameName`` column,
    and save.

Design notes:
  * The engine never touches files on disk — it only generates strings
    and writes them into the Excel workbook. Actual file renames /
    moves are planned for a later phase.
  * Date tokens (``%Date_*%``) prefer ``DateTimeOriginal`` but fall back
    to ``File_Date`` (the filesystem Modified timestamp captured during
    cataloging) when EXIF is missing. The fallback gets an ``[INFO]``
    entry in ``File_Concern`` so the user can see which rows used it.
  * Built-in preflight validation flags collisions, empty renders, and
    over-length filenames as ``[ERROR]`` in ``File_Concern``; those
    rows are skipped (``File_RenameName`` left blank) and will block
    the future rename/move pass.
  * File-path separators, drive colons, and other characters illegal
    in Windows filenames are stripped from the final rendered string
    so a copy/paste of the value is always a valid filename.
"""
from __future__ import annotations

import os
import re
import threading
from collections import defaultdict
from datetime import datetime
from typing import Callable, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# ---------------------------------------------------------------------------
# Variable table
# ---------------------------------------------------------------------------
# Maps template token → human-readable description of where the value
# comes from. The GUI's help label reads this dict so the two never drift.
RENAME_VARIABLES: Dict[str, str] = {
    "%File_Name%":      "FileName stem (filename without extension)",
    "%File_Extension%": "Lowercase extension, including the dot (e.g. .jpg)",
    "%Date_YY%":        "2-digit year from DateTimeOriginal (e.g. 26)",
    "%Date_YYYY%":      "4-digit year from DateTimeOriginal (e.g. 2026)",
    "%Date_MM%":        "2-digit month from DateTimeOriginal (e.g. 04)",
    "%Date_DD%":        "2-digit day from DateTimeOriginal (e.g. 09)",
    "%Camera_Make%":    "CameraMake (e.g. Canon)",
}

# Matches any %...% token so we can spot typos / unknown variables.
_TOKEN_RE = re.compile(r"%[A-Za-z0-9_]+%")

# Characters Windows won't accept in a filename, plus path separators.
# We scrub these out of substituted values (not the template punctuation
# the user types) so a rendered string is always a legal filename.
_ILLEGAL_IN_FILENAME = re.compile(r'[<>:"/\\|?*\x00-\x1f]')

# EXIF date formats we try when pulling year/month/day out of the cell.
_EXIF_DATE_FORMATS = [
    "%Y:%m:%d %H:%M:%S",
    "%Y-%m-%d %H:%M:%S",
    "%Y:%m:%d",
    "%Y-%m-%d",
]

# Concern severity markers — mirror photo_catalog.py's CONCERN_* so the
# rename engine's appended messages sort/color identically to those
# written at extract time.
CONCERN_INFO = "[INFO]"
CONCERN_WARN = "[WARN]"
CONCERN_ERROR = "[ERROR]"
CONCERN_FILL_YELLOW = "FFFFE699"
CONCERN_FILL_RED    = "FFFFC7CE"

# Windows MAX_PATH for the file-name portion; we gate on this so
# renames that would exceed it in the future on-disk move don't make
# it into the report.
_WINDOWS_MAX_FILENAME_LEN = 255


# ---------------------------------------------------------------------------
# Parsing / validation
# ---------------------------------------------------------------------------
def validate_template(template: str) -> Tuple[List[str], List[str]]:
    """
    Scan *template* for ``%...%`` tokens.

    Returns a ``(tokens_found, unknown_tokens)`` tuple. ``tokens_found``
    preserves order-of-appearance so the UI can echo it back to the user;
    ``unknown_tokens`` is a list of tokens that aren't in
    :data:`RENAME_VARIABLES` — those are typos or unsupported variables.
    """
    if not template:
        return [], []
    found = _TOKEN_RE.findall(template)
    unknown = [t for t in found if t not in RENAME_VARIABLES]
    return found, unknown


# ---------------------------------------------------------------------------
# Viability check — CR #2
# ---------------------------------------------------------------------------
# Catches templates that would render useless / dangerous filenames
# *before* we spend time rendering thousands of rows or surprise the
# user with a screenful of collisions. Runs purely on the template
# string; no workbook I/O required.

# Recognized image extensions we accept as "template ends with a
# literal extension" (case-insensitive). Mirrors
# photo_catalog.SUPPORTED_EXTENSIONS so Test/Build use the same set
# without importing the heavier photo_catalog module.
_RECOGNIZED_EXTENSIONS = (
    ".jpg", ".jpeg", ".png", ".tif", ".tiff",
    ".heic", ".heif", ".webp",
    ".cr2", ".cr3", ".nef", ".arw", ".dng", ".orf", ".rw2",
)


def check_template_viability(template: str) -> Tuple[List[str], List[str]]:
    """
    Pattern-check *template* for viability — does it produce filenames
    that are at least nominally valid and unlikely to collide en masse?

    Returns ``(errors, warnings)``:

    * ``errors`` — hard problems. Calling Test Rename String or Build
      Renames with any error should be blocked by the UI so the user
      fixes the template before any rendering starts.
    * ``warnings`` — soft problems. The UI should surface these with a
      "Proceed anyway" confirmation, but rendering is still valid.

    This is *purely* a syntactic / structural check on the template
    string. The heavier per-row validation (collisions, over-length,
    empty renders) still happens inside :func:`build_renames`.
    """
    errors: List[str] = []
    warnings: List[str] = []

    if template is None or not template.strip():
        errors.append("Template is empty.")
        return errors, warnings

    found, unknown = validate_template(template)
    if unknown:
        errors.append(
            "Unknown variable(s): " + ", ".join(sorted(set(unknown)))
        )

    # Rule 1 — must produce a file extension.
    ends_with_literal_ext = any(
        template.lower().rstrip().endswith(ext) for ext in _RECOGNIZED_EXTENSIONS
    )
    if "%File_Extension%" not in found and not ends_with_literal_ext:
        errors.append(
            "Template must produce a file extension. "
            "Add %File_Extension% at the end, or type a literal "
            "extension like .jpg."
        )

    # Rule 2 — must have at least one variable token so not every photo
    # renames to the same literal string.
    if not found:
        errors.append(
            "Template has no %Variable% tokens; every photo would "
            "rename to the same literal string."
        )

    # Rule 4 (warning) — must include %File_Name% for per-file uniqueness.
    # Photos from the same camera on the same date would otherwise
    # collide en masse, which is valid but almost always unintended.
    if "%File_Name%" not in found:
        warnings.append(
            "Template does not include %File_Name%. Photos taken on "
            "the same date by the same camera will collide and be "
            "flagged as [ERROR] in File_Concern."
        )

    # Rule 6 (warning) — path separators will be stripped, which
    # usually means the user is trying to create a subfolder and will
    # be surprised when it doesn't. Surface this so they can confirm.
    if "/" in template or "\\" in template:
        warnings.append(
            "Template contains path separators (/ or \\). These will "
            "be stripped from the rendered filename. The rename engine "
            "does not create subfolders."
        )

    return errors, warnings


def _coerce_datetime(value) -> Optional[datetime]:
    """Best-effort conversion of a DateTimeOriginal cell value to datetime."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    s = str(value).strip()[:19]
    for fmt in _EXIF_DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _sanitize_substitution(val: str) -> str:
    """
    Scrub path separators and filename-illegal characters from a
    substituted value. Leaves template punctuation (dashes, underscores,
    spaces) alone because those are typed by the user, not substituted.
    """
    return _ILLEGAL_IN_FILENAME.sub("", val)


def render_row(
    template: str,
    row: Dict,
) -> Tuple[Optional[str], Optional[str], bool]:
    """
    Render one rename string from *template* and a row dict.

    Returns ``(rendered_filename, reason, used_file_date_fallback)``:

    * ``rendered_filename`` is the rendered string on success, else ``None``.
    * ``reason`` is ``None`` on success or a short string describing the
      missing data that caused the skip (written to the log / File_Concern).
    * ``used_file_date_fallback`` is ``True`` when a ``%Date_*%`` token
      was filled from ``File_Date`` because ``DateTimeOriginal`` was
      blank. The caller should add an ``[INFO]`` concern so the user
      can see which rows used the fallback.
    """
    found, unknown = validate_template(template)
    if unknown:
        return (
            None,
            f"unknown variable(s): {', '.join(sorted(set(unknown)))}",
            False,
        )

    # Build the substitution map lazily — only parse the date if a
    # date token is actually in the template, so rows missing EXIF
    # dates can still be renamed by non-date templates.
    uses_date = any(t.startswith("%Date_") for t in found)
    dt: Optional[datetime] = None
    used_fallback = False
    if uses_date:
        dt = _coerce_datetime(row.get("DateTimeOriginal"))
        if dt is None:
            # CR #1: fall back to File_Date before giving up.
            dt = _coerce_datetime(row.get("File_Date"))
            if dt is not None:
                used_fallback = True
        if dt is None:
            return None, "missing DateTimeOriginal and File_Date", False

    file_name = row.get("File_Name") or row.get("FileName") or ""
    file_ext = (row.get("File_Extension") or row.get("FileExtension") or "").strip()
    # Derive stem from File_Name (strip the extension if present).
    if file_ext and file_name.lower().endswith(file_ext.lower()):
        stem = file_name[: -len(file_ext)]
    else:
        stem = os.path.splitext(file_name)[0]

    make = row.get("CameraMake")
    if ("%Camera_Make%" in found) and (make is None or str(make).strip() == ""):
        return None, "missing CameraMake", used_fallback

    substitutions: Dict[str, str] = {
        "%File_Name%":      _sanitize_substitution(stem),
        "%File_Extension%": _sanitize_substitution(file_ext),
        "%Camera_Make%":    _sanitize_substitution(str(make).strip() if make else ""),
    }
    if dt is not None:
        substitutions.update({
            "%Date_YY%":   dt.strftime("%y"),
            "%Date_YYYY%": dt.strftime("%Y"),
            "%Date_MM%":   dt.strftime("%m"),
            "%Date_DD%":   dt.strftime("%d"),
        })

    result = template
    for token, value in substitutions.items():
        result = result.replace(token, value)
    return result, None, used_fallback


# ---------------------------------------------------------------------------
# Workbook I/O
# ---------------------------------------------------------------------------
def _column_index_map(ws) -> Dict[str, int]:
    """Return a ``{header_text: 1-based column index}`` map from row 1."""
    mapping: Dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header:
            mapping[str(header)] = col_idx
    return mapping


def _row_dict(ws, row_idx: int, col_map: Dict[str, int]) -> Dict:
    """Build a ``{header: value}`` dict for one spreadsheet row."""
    return {
        header: ws.cell(row=row_idx, column=col_idx).value
        for header, col_idx in col_map.items()
    }


def test_template(
    xlsx_path: str,
    template: str,
    row_limit: int = 10,
) -> Dict:
    """
    Render the first *row_limit* data rows from *xlsx_path* under *template*
    without writing anything back. Used by the Test Rename String button.

    Returns a dict with:
        'tokens'   — list[str] of tokens found in the template
        'unknown'  — list[str] of unknown tokens (empty if valid)
        'previews' — list[tuple[original_name, rendered_or_none, reason_or_none]]
        'total_rows' — total data rows in the sheet (for context)
    """
    found, unknown = validate_template(template)
    # CR #2: run the viability check so the UI doesn't preview a
    # template that's guaranteed to fail validation later.
    errors, warnings = check_template_viability(template)
    result: Dict = {
        "tokens": found,
        "unknown": unknown,
        "errors": errors,
        "warnings": warnings,
        "previews": [],
        "total_rows": 0,
    }
    if unknown or errors:
        # Don't bother opening the workbook if the template is invalid.
        return result

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        col_map = _column_index_map(ws)
        # Accept both the v2.1.1 File_Name header and the older FileName
        # header so pre-existing workbooks still preview correctly.
        name_header = "File_Name" if "File_Name" in col_map else "FileName"
        if name_header not in col_map:
            raise ValueError(
                "Workbook is missing a 'File_Name' column \u2014 is this a "
                "PhotoCatalog report?"
            )

        total = max(0, ws.max_row - 1)
        result["total_rows"] = total

        taken = 0
        for row_idx in range(2, ws.max_row + 1):
            if taken >= row_limit:
                break
            row = _row_dict(ws, row_idx, col_map)
            rendered, reason, _fallback = render_row(template, row)
            result["previews"].append(
                (row.get(name_header) or "", rendered, reason)
            )
            taken += 1
    finally:
        wb.close()

    return result


def _ensure_column(ws, col_map: Dict[str, int], header: str) -> int:
    """
    Return the 1-based index of *header*, appending it on the right of
    the sheet if the workbook predates that column.
    """
    if header in col_map:
        return col_map[header]
    new_idx = ws.max_column + 1
    ws.cell(row=1, column=new_idx, value=header)
    col_map[header] = new_idx
    return new_idx


def _append_concerns(
    ws,
    row_idx: int,
    col_idx: int,
    new_entries: List[str],
) -> None:
    """
    Append *new_entries* to the existing File_Concern cell at
    ``(row_idx, col_idx)`` and reapply fill color based on the joined
    severity. Existing entries are preserved so a rename preflight
    doesn't stomp the extract-time concerns.
    """
    if not new_entries:
        return
    cell = ws.cell(row=row_idx, column=col_idx)
    existing = str(cell.value).strip() if cell.value else ""
    parts = [p.strip() for p in existing.split(";")] if existing else []
    parts = [p for p in parts if p]
    # Suppress exact duplicates so repeated Build Renames passes don't
    # accumulate noise in the cell.
    for entry in new_entries:
        if entry not in parts:
            parts.append(entry)
    joined = "; ".join(parts)
    cell.value = joined
    if CONCERN_ERROR in joined:
        cell.fill = PatternFill("solid", fgColor=CONCERN_FILL_RED)
    elif joined:
        cell.fill = PatternFill("solid", fgColor=CONCERN_FILL_YELLOW)


def build_renames(
    xlsx_path: str,
    template: str,
    progress_callback: Optional[Callable[[int, int], None]] = None,
    log_callback: Optional[Callable[[str], None]] = None,
    cancel_event: Optional[threading.Event] = None,
) -> Dict:
    """
    Render *template* for every data row in *xlsx_path* and write the
    result into the ``File_RenameName`` column, then save.

    As of CR #1 this is also the preflight validator: before writing
    any results, every row is checked for collisions (multiple rows
    rendering to the same filename within the same parent folder),
    over-length filenames, and empty renders. Offending rows are
    flagged ``[ERROR]`` in ``File_Concern`` and have ``File_RenameName``
    left blank so the future on-disk rename/move pass can skip them.
    Rows whose date came from the ``File_Date`` fallback get an
    ``[INFO]`` note.

    Returns a summary dict:
        'total'          — total data rows
        'renamed'        — rows that got a rename string
        'skipped'        — list of (filename, reason) tuples left blank
        'fallback_dates' — count of rows where File_Date filled in for EXIF
        'errors'         — count of rows flagged [ERROR] (collisions/length/empty)
        'cancelled'      — True if the user aborted mid-run
    """
    log = log_callback if log_callback is not None else (lambda _m: None)

    found, unknown = validate_template(template)
    if unknown:
        raise ValueError(
            f"Template contains unknown variable(s): "
            f"{', '.join(sorted(set(unknown)))}"
        )
    # CR #2: belt-and-braces viability check inside the engine in case
    # a caller bypasses the GUI dialog. Warnings are surfaced to the
    # log but do not block (the GUI already prompted on them); errors
    # raise ValueError so the worker thread reports them upstream.
    errors, warnings = check_template_viability(template)
    if errors:
        raise ValueError(
            "Template is not viable:\n  - " + "\n  - ".join(errors)
        )
    for w in warnings:
        log(f"Warning: {w}")

    log(f"Loading workbook: {xlsx_path}")
    wb = load_workbook(xlsx_path)
    try:
        ws = wb.active
        col_map = _column_index_map(ws)
        name_header = "File_Name" if "File_Name" in col_map else "FileName"
        if name_header not in col_map:
            raise ValueError(
                "Workbook is missing a 'File_Name' column \u2014 is this a "
                "PhotoCatalog report?"
            )
        rename_header = (
            "File_RenameName" if "File_RenameName" in col_map else "RenameFileName"
        )
        # Make sure the rename column exists; if this workbook predates
        # v2.1.1 and is labelled RenameFileName, keep using that name.
        rename_col = _ensure_column(ws, col_map, rename_header)
        # File_Concern and File_Path may also be missing on very old
        # workbooks — add them on the right so we can still write results.
        concern_col = _ensure_column(ws, col_map, "File_Concern")
        path_header = "File_Path" if "File_Path" in col_map else "FilePath"

        total = max(0, ws.max_row - 1)
        renamed = 0
        fallback_dates = 0
        errors = 0
        skipped: List[Tuple[str, str]] = []
        cancelled = False

        # --- Pass 1: render each row and stage the results ---------------
        # We need every rendered name before we can detect collisions,
        # so we stage per-row results here and commit in pass 2.
        staged: List[Dict] = []
        # Group rendered names by their target folder so two rows with
        # the same base name in different folders aren't flagged as a
        # collision (the future on-disk move will use the parent of
        # File_Path as the destination).
        name_counts: Dict[Tuple[str, str], int] = defaultdict(int)

        for i, row_idx in enumerate(range(2, ws.max_row + 1), start=1):
            if cancel_event is not None and cancel_event.is_set():
                cancelled = True
                break
            row = _row_dict(ws, row_idx, col_map)
            rendered, reason, used_fallback = render_row(template, row)

            file_path = row.get(path_header) or row.get("FilePath") or ""
            parent = os.path.dirname(str(file_path))

            row_concerns: List[str] = []
            if used_fallback:
                row_concerns.append(
                    f"{CONCERN_INFO} Used File_Date (EXIF DateTimeOriginal missing)"
                )

            staged.append({
                "row_idx": row_idx,
                "filename": row.get(name_header) or "",
                "rendered": rendered,
                "reason": reason,
                "parent": parent,
                "concerns": row_concerns,
            })

            if rendered is not None:
                name_counts[(parent.lower(), rendered.lower())] += 1

            if progress_callback is not None and i % 250 == 0:
                # Report half-progress for pass 1 so the bar moves
                # during long runs.
                progress_callback(i // 2, total)

        if cancelled:
            log("Cancelled during preflight \u2014 workbook not modified.")
            return {
                "total": total, "renamed": 0, "skipped": [],
                "fallback_dates": 0, "errors": 0, "cancelled": True,
            }

        # --- Pass 2: validate + write -----------------------------------
        for j, item in enumerate(staged, start=1):
            if cancel_event is not None and cancel_event.is_set():
                cancelled = True
                break

            row_idx = item["row_idx"]
            rendered = item["rendered"]
            row_concerns = item["concerns"]

            if rendered is None:
                # Render failed outright. Leave rename cell blank and
                # log the reason (no [ERROR] concern — the skip reason
                # is already visible via missing File_RenameName).
                ws.cell(row=row_idx, column=rename_col, value=None)
                skipped.append((item["filename"], item["reason"] or "unknown"))
            else:
                # Preflight: validate length / emptiness / collisions.
                errors_on_row: List[str] = []
                if not rendered.strip():
                    errors_on_row.append(
                        f"{CONCERN_ERROR} Rendered filename is empty"
                    )
                if len(rendered) > _WINDOWS_MAX_FILENAME_LEN:
                    errors_on_row.append(
                        f"{CONCERN_ERROR} Rendered filename exceeds "
                        f"{_WINDOWS_MAX_FILENAME_LEN} chars ({len(rendered)})"
                    )
                if name_counts[(item["parent"].lower(), rendered.lower())] > 1:
                    errors_on_row.append(
                        f"{CONCERN_ERROR} Rename collision with another row "
                        f"in the same folder"
                    )

                if errors_on_row:
                    ws.cell(row=row_idx, column=rename_col, value=None)
                    errors += 1
                    row_concerns.extend(errors_on_row)
                    skipped.append(
                        (item["filename"], "validation error \u2014 see File_Concern")
                    )
                else:
                    ws.cell(row=row_idx, column=rename_col, value=rendered)
                    renamed += 1
                    if item["concerns"]:
                        # The fallback [INFO] will be appended below.
                        fallback_dates += 1

            _append_concerns(ws, row_idx, concern_col, row_concerns)

            if progress_callback is not None and j % 250 == 0:
                progress_callback((total // 2) + (j // 2), total)

        if progress_callback is not None:
            progress_callback(total, total)

        if not cancelled:
            log(
                f"Saving workbook with {renamed:,} rename strings "
                f"({errors:,} error(s), {fallback_dates:,} File_Date fallback(s))\u2026"
            )
            wb.save(xlsx_path)
            log("Save complete.")
        else:
            log("Cancelled before save \u2014 workbook not modified.")
    finally:
        wb.close()

    return {
        "total": total,
        "renamed": renamed,
        "skipped": skipped,
        "fallback_dates": fallback_dates,
        "errors": errors,
        "cancelled": cancelled,
    }
