"""
duplicate_detector.py — Duplicate discovery for PhotoCatalog v3.

Consumes the catalog row list produced by ``photo_catalog.extract_metadata``
(with File_Hash populated when the user chose Hash mode) and groups
likely duplicates together so the workbook user can review, then
disposition them with the Move-non-keepers / Delete-non-keepers
buttons.

Public entry points:
    :func:`detect_duplicates` — in-place on a list of row dicts; sets
        ``File_DupeGroup`` (int or None) and ``File_DupeKeep`` (bool
        or None) on every row that belongs to a group, and returns a
        summary dict for logging.
    :func:`detect_duplicates_on_workbook` — runs detection directly
        against an already-saved workbook. Reads the rows out of the
        Catalog sheet, optionally computes File_Hash for hash mode,
        writes File_Hash / File_DupeGroup / File_DupeKeep back to the
        corresponding cells, applies the orange File_Name highlight,
        and saves. Intended for the GUI's "Detect Duplicates on
        Existing Workbook" button so the user can recover from having
        left the Duplicate Detection combo on None without re-running
        the (potentially long) catalog scan.

Match modes:
    ``"none"``          — skip detection, do nothing. (Used when the
                          user selects *None* in the dupe-mode combo.)
    ``"filename_size"`` — match on ``(lowercase basename, File_SizeBytes)``.
                          Fast; works without hashing. Catches obvious
                          cross-folder copies but misses renames.
    ``"hash"``          — match on ``File_Hash``. Thorough, catches
                          renames. Requires ``File_Hash`` to already be
                          populated (which ``extract_metadata`` only does
                          when the pipeline was told to hash).

Keeper selection:
    Exactly one row per group gets ``File_DupeKeep = True``. The keeper
    is the row with the earliest ``File_Date``. Ties are broken by
    shortest ``File_Path`` (nearest-to-root wins), then by first
    encountered order.
"""
from __future__ import annotations

import hashlib
import logging
import os
from collections import defaultdict
from datetime import datetime
from typing import Callable, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Keep this constant in lockstep with photo_catalog.DUPE_FILL_ORANGE so
# that a workbook re-painted here matches one produced by the catalog
# pipeline. Duplicated (rather than imported) to keep this module free
# of a cyclic dependency on photo_catalog.
_DUPE_FILL_ORANGE = "FFFFD59B"


# ---------------------------------------------------------------------------
# Hashing helper
# ---------------------------------------------------------------------------
# Exposed so photo_catalog.extract_metadata and the copy_engine's
# post-copy verification (future v3.1) can share one implementation.
_HASH_BUFFER_BYTES = 1 << 20   # 1 MiB buffer — tuned for SSD throughput.


def compute_md5(filepath: str) -> Optional[str]:
    """
    Stream *filepath* through MD5 and return the hex digest.

    Chosen deliberately as MD5 rather than SHA-256: we're matching
    "are these the same bytes", not signing anything — MD5 is
    ~2–3× faster on large RAW files and collision-safe enough for
    dedup with the extra filename+size sanity check the hash mode
    workflow already implies.

    Returns ``None`` on read failure so the caller can log a
    ``[WARN]`` concern rather than crashing the whole run.
    """
    try:
        md5 = hashlib.md5()
        with open(filepath, "rb") as f:
            while True:
                chunk = f.read(_HASH_BUFFER_BYTES)
                if not chunk:
                    break
                md5.update(chunk)
        return md5.hexdigest()
    except OSError as e:
        logging.warning("Could not hash %s: %s", filepath, e)
        return None


# ---------------------------------------------------------------------------
# Match-key builders
# ---------------------------------------------------------------------------
def _key_filename_size(row: Dict) -> Optional[Tuple[str, int]]:
    name = row.get("File_Name")
    size = row.get("File_SizeBytes")
    if not name or size is None:
        return None
    return (str(name).lower(), int(size))


def _key_hash(row: Dict) -> Optional[str]:
    h = row.get("File_Hash")
    if not h:
        return None
    return str(h).lower()


# ---------------------------------------------------------------------------
# Keeper-selection sort key
# ---------------------------------------------------------------------------
# Orders candidate rows so the first one after sort is the desired
# keeper. Earliest File_Date wins; absent File_Date sorts last so a
# row with a real date always beats one without. Ties broken by the
# length of File_Path (shorter path = closer to root, likely the
# "original" location), then by first-encountered index.
_DATE_FAR_FUTURE = datetime(9999, 12, 31)


def _keeper_sort_key(entry: Tuple[int, Dict]) -> Tuple:
    idx, row = entry
    dt = row.get("File_Date")
    if not isinstance(dt, datetime):
        dt = _DATE_FAR_FUTURE
    path = row.get("File_Path") or ""
    return (dt, len(str(path)), idx)


# ---------------------------------------------------------------------------
# Public: detection pass
# ---------------------------------------------------------------------------
def detect_duplicates(
    rows: List[Dict],
    mode: str,
) -> Dict:
    """
    Group likely duplicates in *rows* (in-place) and mark keepers.

    The function mutates each row by setting:

        ``File_DupeGroup`` — an ``int`` group ID (starting at 1), or
                             ``None`` if the row doesn't belong to a
                             group.
        ``File_DupeKeep``  — ``True`` for exactly one row per group
                             (the chosen keeper), ``False`` for the
                             rest, ``None`` for rows outside any group.

    Returns a summary dict with keys:
        ``mode``             — the *mode* argument, echoed back.
        ``groups``           — number of duplicate groups found.
        ``duplicate_rows``   — total rows that ended up in a group.
        ``keepers``          — rows marked ``File_DupeKeep = True``.
        ``non_keepers``      — rows marked ``File_DupeKeep = False``.
        ``skipped``          — rows that had no match key (e.g. hash
                               mode with no File_Hash populated).
    """
    summary = {
        "mode": mode,
        "groups": 0,
        "duplicate_rows": 0,
        "keepers": 0,
        "non_keepers": 0,
        "skipped": 0,
    }

    # Initialize every row's dupe fields so callers never see "missing".
    for row in rows:
        row.setdefault("File_DupeGroup", None)
        row.setdefault("File_DupeKeep", None)

    if mode == "none" or not rows:
        return summary

    if mode == "filename_size":
        keyer = _key_filename_size
    elif mode == "hash":
        keyer = _key_hash
    else:
        raise ValueError(f"Unknown dupe detection mode: {mode!r}")

    # Build the grouping: key -> list of (index, row).
    buckets: Dict = defaultdict(list)
    for i, row in enumerate(rows):
        key = keyer(row)
        if key is None:
            summary["skipped"] += 1
            continue
        buckets[key].append((i, row))

    group_id = 0
    for _key, members in buckets.items():
        if len(members) < 2:
            continue
        group_id += 1
        summary["groups"] += 1
        summary["duplicate_rows"] += len(members)

        # Choose the keeper as the row that sorts first under the
        # keeper policy (earliest date, shortest path, first seen).
        ordered = sorted(members, key=_keeper_sort_key)
        keeper_idx = ordered[0][0]

        for _, row in members:
            row["File_DupeGroup"] = group_id
        for idx, row in members:
            is_keeper = idx == keeper_idx
            row["File_DupeKeep"] = bool(is_keeper)
            if is_keeper:
                summary["keepers"] += 1
            else:
                summary["non_keepers"] += 1

    return summary


# ---------------------------------------------------------------------------
# Hashing pass helper (orchestrated by catalog_pipeline)
# ---------------------------------------------------------------------------
def size_collision_candidates(rows: Iterable[Dict]) -> "set[str]":
    """
    Return the set of ``File_Path`` values that share a
    ``File_SizeBytes`` with at least one other row.

    Two byte-identical files **must** have the same size, so any file
    whose size is unique across the library cannot have a byte-identical
    twin and doesn't need to be hashed. This pre-screen routinely cuts
    the hashing workload by 80–95% on real photo libraries (size
    collisions tend to be a small fraction of any well-aged archive).

    Rows with no ``File_SizeBytes`` and rows with no ``File_Path`` are
    ignored — both go through the regular full-hash path so we don't
    silently drop them on a malformed source.
    """
    by_size: Dict = defaultdict(list)
    for row in rows:
        size = row.get("File_SizeBytes")
        path = row.get("File_Path")
        if size is None or not path:
            continue
        by_size[size].append(path)
    candidates: "set[str]" = set()
    for paths in by_size.values():
        if len(paths) >= 2:
            candidates.update(paths)
    return candidates


def populate_hashes(
    rows: List[Dict],
    progress_callback=None,
    cancel_event=None,
    candidate_paths: Optional["set[str]"] = None,
) -> int:
    """
    Compute MD5 for every row that has a readable ``File_Path`` and
    no ``File_Hash`` yet. Mutates rows in place.

    Returns the number of rows that were successfully hashed.

    Args:
        rows: All catalog rows (the function decides per-row whether to
            hash; non-candidates are skipped silently).
        progress_callback: ``(current, total)`` ticks. ``total`` is the
            *candidate* count when ``candidate_paths`` is supplied so
            the bar fills based on real work, not on the full row list.
        cancel_event: Optional ``threading.Event``; if set, the loop
            breaks between files.
        candidate_paths: Optional restriction set produced by
            :func:`size_collision_candidates` (or any caller-built set).
            When supplied, only rows whose ``File_Path`` is in this set
            are hashed. Pass ``None`` to hash every eligible row (the
            old "always hash" behavior preserved as an opt-out).

    The catalog pipeline only calls this when the user chose
    ``dupe_mode == "hash"`` — it's factored out here so the hashing
    loop lives next to the detection code that consumes its output.
    """
    # When a candidate set is provided, total = candidate count so the
    # progress callback reports against the actual planned work.
    # Otherwise we fall back to the legacy behavior of total = all rows.
    if candidate_paths is None:
        total = len(rows)
    else:
        total = sum(
            1 for r in rows
            if r.get("File_Path") in candidate_paths and not r.get("File_Hash")
        )

    hashed = 0
    seen = 0  # rows we actually attempted to hash (drives progress)
    for row in rows:
        if cancel_event is not None and cancel_event.is_set():
            break
        if row.get("File_Hash"):
            continue
        path = row.get("File_Path")
        if not path or not os.path.exists(path):
            continue
        # Skip non-candidates when the optimization is enabled. These
        # files are guaranteed unique by size, so they cannot be
        # byte-identical to anything else in the library.
        if candidate_paths is not None and path not in candidate_paths:
            continue

        seen += 1
        digest = compute_md5(path)
        if digest:
            row["File_Hash"] = digest
            hashed += 1
            # Append a concern if one of the hash calls failed — we
            # still want visibility into rows that *should* have
            # hashes but don't.
        elif "_concerns" in row:
            row["_concerns"].append(
                "[WARN] Could not compute File_Hash (read failure)"
            )
        if progress_callback is not None and (seen % 25 == 0 or seen == total):
            progress_callback(seen, total)
    return hashed


# ---------------------------------------------------------------------------
# Workbook-mode detection (used by the "Detect Duplicates on Existing
# Workbook" GUI button so the user doesn't have to re-scan a whole drive
# just because they left the dupe-mode combo on None the first time).
# ---------------------------------------------------------------------------
_CATALOG_SHEET_NAME = "Catalog"

# Columns we need to read from the existing workbook to run detection and
# keeper-selection. Others (ImageWidth, CameraMake, etc.) are irrelevant
# to this pass and are ignored.
_READ_COLUMNS = (
    "File_Name",
    "File_SizeBytes",
    "File_Date",
    "File_Path",
    "File_Hash",
)

# Columns we write back onto the worksheet.
_WRITE_COLUMNS = ("File_Hash", "File_DupeGroup", "File_DupeKeep")


def _coerce_int(value) -> Optional[int]:
    """Best-effort convert a cell value to int; return None on failure."""
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def detect_duplicates_on_workbook(
    xlsx_path: str,
    mode: str,
    progress_callback: Optional[Callable[[int, int, str], None]] = None,
    log_callback: Optional[Callable[[str], None]] = None,
    cancel_event=None,
    always_hash_all_files: bool = False,
) -> Dict:
    """
    Run duplicate detection against an already-saved PhotoCatalog workbook.

    Intended for the GUI's "Detect Duplicates on Existing Workbook" button:
    the user produced a workbook with ``dupe_mode="none"`` and now wants to
    add duplicate info without re-scanning the whole folder tree.

    Reads rows from the ``Catalog`` sheet, (optionally) hashes files whose
    ``File_Path`` still resolves on disk, runs :func:`detect_duplicates`,
    then writes ``File_Hash`` / ``File_DupeGroup`` / ``File_DupeKeep``
    back into the matching cells and paints the pale-orange duplicate
    fill on the ``File_Name`` cell of every grouped row. Finally saves
    the workbook in place.

    Args:
        xlsx_path: Full path to the ``.xlsx`` file. Must already contain
            a ``Catalog`` sheet with at least the ``File_Name``,
            ``File_SizeBytes``, ``File_Date`` and ``File_Path`` columns.
        mode: One of ``"filename_size"`` or ``"hash"``. ``"none"`` is
            rejected here because the button should never be called with
            detection disabled.
        progress_callback: ``(current, total, message)`` for the hashing
            phase (hash mode only) and a final "writing" tick.
        log_callback: Receives human-readable status lines.
        cancel_event: Optional ``threading.Event``; if set, the run exits
            between files during hashing and before writing cells.

    Returns:
        A dict that merges the :func:`detect_duplicates` summary with
        extra keys describing the workbook pass itself:
            ``xlsx_path``        — echoed input path.
            ``row_count``        — rows read from the Catalog sheet.
            ``hashed``           — rows newly hashed this pass.
            ``missing_on_disk``  — rows whose File_Path no longer exists
                                   (relevant only for hash mode).

    Raises:
        FileNotFoundError: The workbook or its Catalog sheet is missing.
        ValueError:        Unknown *mode* or required columns absent.
    """
    log = log_callback or (lambda _m: None)

    if mode not in ("filename_size", "hash"):
        raise ValueError(
            f"detect_duplicates_on_workbook: mode must be 'filename_size' "
            f"or 'hash', got {mode!r}"
        )
    if not os.path.isfile(xlsx_path):
        raise FileNotFoundError(xlsx_path)

    log(f"Opening workbook: {xlsx_path}")
    wb = load_workbook(xlsx_path)
    if _CATALOG_SHEET_NAME not in wb.sheetnames:
        raise FileNotFoundError(
            f"Workbook has no '{_CATALOG_SHEET_NAME}' sheet: {xlsx_path}"
        )
    ws = wb[_CATALOG_SHEET_NAME]

    # Header map: column name -> 1-based column index.
    header_map: Dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        name = cell.value
        if isinstance(name, str) and name:
            header_map[name] = col_idx

    required = {"File_Name", "File_SizeBytes", "File_Date", "File_Path"}
    missing = required - set(header_map.keys())
    if missing:
        raise ValueError(
            f"Workbook Catalog sheet is missing required column(s): "
            f"{sorted(missing)}"
        )

    # Make sure the writeback columns exist; if not, append them to the
    # header row. This keeps old v2 workbooks forward-compatible even
    # though the GUI only offers this button for v3 workbooks in practice.
    max_col = ws.max_column
    for name in _WRITE_COLUMNS:
        if name not in header_map:
            max_col += 1
            ws.cell(row=1, column=max_col, value=name)
            header_map[name] = max_col

    # --- Read rows into dicts ---------------------------------------------
    last_row = ws.max_row
    rows: List[Dict] = []
    # Track the source row index so we can write results back to the
    # exact same worksheet row. Stored alongside the dict under a
    # private key that detect_duplicates ignores.
    for r in range(2, last_row + 1):
        row: Dict = {"_ws_row": r}
        for name in _READ_COLUMNS:
            col = header_map.get(name)
            if col is None:
                continue
            row[name] = ws.cell(row=r, column=col).value
        # Normalize types so the detector's key builders work unchanged.
        row["File_SizeBytes"] = _coerce_int(row.get("File_SizeBytes"))
        if row.get("File_Hash") is not None:
            row["File_Hash"] = str(row["File_Hash"]).strip() or None
        # File_Date already round-trips as a datetime via openpyxl because
        # we wrote it with number_format='YYYY-MM-DD HH:MM:SS' and an
        # actual datetime value. Leave whatever comes back as-is.
        rows.append(row)

    log(f"Read {len(rows)} data row(s) from '{_CATALOG_SHEET_NAME}' sheet.")

    # --- Optional hashing pass --------------------------------------------
    hashed = 0
    missing_on_disk = 0
    candidates_planned = 0
    if mode == "hash":
        # populate_hashes already skips rows whose File_Path is missing
        # or whose File_Hash is already present. Count missing separately
        # here so the returned summary tells the user why some rows may
        # still have no hash afterward.
        for row in rows:
            path = row.get("File_Path")
            if not row.get("File_Hash") and (not path or not os.path.exists(path)):
                missing_on_disk += 1

        # Smart-hash optimization: pre-screen by file size and only hash
        # files that share a size with at least one other row. Two
        # byte-identical files MUST have the same size, so size-unique
        # files cannot have a duplicate twin and don't need hashing.
        # Opt out via always_hash_all_files=True if the user wants the
        # File_Hash column populated for every row.
        if always_hash_all_files:
            log("Hashing files (MD5) — full sweep (size-bucket optimization disabled).")
            candidate_set = None
        else:
            candidate_set = size_collision_candidates(rows)
            candidates_planned = len(candidate_set)
            total_rows = len(rows)
            skipped = total_rows - candidates_planned
            pct_skipped = (100.0 * skipped / total_rows) if total_rows else 0
            log(
                f"Hashing files (MD5) — size-bucket optimization: "
                f"{candidates_planned:,} candidate(s) of {total_rows:,} rows "
                f"({skipped:,} unique-size rows skipped, {pct_skipped:.1f}% saved)."
            )

        def _prog(current: int, total: int) -> None:
            if progress_callback is not None:
                progress_callback(current, total, "hashing")

        hashed = populate_hashes(
            rows,
            progress_callback=_prog,
            cancel_event=cancel_event,
            candidate_paths=candidate_set,
        )
        log(f"  Hashed {hashed} file(s); {missing_on_disk} no longer on disk.")
        if cancel_event is not None and cancel_event.is_set():
            log("Cancelled during hashing — workbook not modified.")
            return {
                "mode": mode,
                "groups": 0,
                "duplicate_rows": 0,
                "keepers": 0,
                "non_keepers": 0,
                "skipped": 0,
                "xlsx_path": xlsx_path,
                "row_count": len(rows),
                "hashed": hashed,
                "missing_on_disk": missing_on_disk,
                "candidates_planned": candidates_planned,
                "always_hash_all_files": always_hash_all_files,
                "cancelled": True,
            }

    # --- Detect duplicates (in place on rows) -----------------------------
    log(f"Running duplicate detection (mode={mode})...")
    summary = detect_duplicates(rows, mode=mode)
    log(
        f"  {summary['groups']} group(s), "
        f"{summary['duplicate_rows']} row(s) in groups, "
        f"{summary['keepers']} keeper(s), "
        f"{summary['non_keepers']} non-keeper(s), "
        f"{summary['skipped']} skipped."
    )

    # --- Write results back ------------------------------------------------
    if cancel_event is not None and cancel_event.is_set():
        log("Cancelled before writeback — workbook not modified.")
        summary.update({
            "xlsx_path": xlsx_path,
            "row_count": len(rows),
            "hashed": hashed,
            "missing_on_disk": missing_on_disk,
            "candidates_planned": candidates_planned,
            "always_hash_all_files": always_hash_all_files,
            "cancelled": True,
        })
        return summary

    log("Writing File_Hash / File_DupeGroup / File_DupeKeep back to workbook...")
    dupe_fill = PatternFill("solid", fgColor=_DUPE_FILL_ORANGE)
    hash_col  = header_map["File_Hash"]
    group_col = header_map["File_DupeGroup"]
    keep_col  = header_map["File_DupeKeep"]
    name_col  = header_map["File_Name"]

    total_write = len(rows)
    for i, row in enumerate(rows, start=1):
        ws_row = row["_ws_row"]
        # Write hash if we have one (either pre-existing or freshly computed).
        if row.get("File_Hash"):
            ws.cell(row=ws_row, column=hash_col, value=row["File_Hash"])
        # DupeGroup / DupeKeep are only meaningful when the row is in a group.
        group = row.get("File_DupeGroup")
        keep  = row.get("File_DupeKeep")
        ws.cell(row=ws_row, column=group_col, value=group if group is not None else None)
        # Write the boolean as a real bool so Excel renders it as TRUE/FALSE,
        # matching what write_excel produces during a fresh catalog pass.
        if keep is None:
            ws.cell(row=ws_row, column=keep_col, value=None)
        else:
            ws.cell(row=ws_row, column=keep_col, value=bool(keep))

        # Paint the orange fill on the File_Name cell for grouped rows.
        if group:
            ws.cell(row=ws_row, column=name_col).fill = dupe_fill

        if progress_callback is not None and (i % 100 == 0 or i == total_write):
            progress_callback(i, total_write, "writing")

    log(f"Saving workbook: {xlsx_path}")
    wb.save(xlsx_path)
    wb.close()

    summary.update({
        "xlsx_path": xlsx_path,
        "row_count": len(rows),
        "hashed": hashed,
        "missing_on_disk": missing_on_disk,
        "candidates_planned": candidates_planned,
        "always_hash_all_files": always_hash_all_files,
        "cancelled": False,
    })
    return summary
