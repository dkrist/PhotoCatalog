"""
copy_engine.py — Copy / Move-non-keepers / Delete-non-keepers / source
cleanup for PhotoCatalog v3.

Drives the filesystem operations that sit between the v2.x catalog
workflow and the v3 "reorganize a messy drive" goal. Reads the
catalog workbook (which already has File_DestFolder / File_DestPath /
File_DupeKeep populated by the extract + dupe-detection + rename
passes) and moves bytes around accordingly, writing a rollback
journal for every destructive-ish operation.

Public entry points:
    :func:`populate_destination_columns` — given a workbook, a rename
        template, and a FolderConfig, fill File_RenameName,
        File_DestFolder, and File_DestPath for every row. Called by
        the catalog pipeline before write_excel.
    :func:`copy_to_destination`          — Copy pass. Reads the workbook,
        copies every row's source to its File_DestPath, follows same-
        stem sidecars, updates File_Status to ``Copied``.
    :func:`move_non_keepers`             — moves rows with
        ``File_DupeKeep = FALSE`` from their source to a user-chosen
        holding folder. Updates File_Status to ``DupeMoved``.
    :func:`delete_non_keepers`           — deletes rows with
        ``File_DupeKeep = FALSE`` from the source drive. Updates
        File_Status to ``DupeDeleted``.
    :func:`find_empty_source_folders`    — walks the source tree,
        lists folders that are now empty (no non-hidden files
        remaining).
    :func:`remove_empty_source_folders`  — deletes folders returned
        by the previous call, journaling each rmdir.

Design notes:
    * All destructive operations accept an optional ``cancel_event``
      so the GUI's Cancel button can halt a long run mid-way.
    * Each operation runs in two passes: workbook read (collect the
      plan) then filesystem pass (execute + journal + write back to
      the workbook). The workbook is only saved at the end so a
      cancel leaves a coherent state.
    * Sidecars follow the primary. Sidecars are detected at filesystem-
      walk time from the primary's parent folder, not cataloged as
      workbook rows.
"""
from __future__ import annotations

import logging
import os
import shutil
import threading
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from folder_composer import FolderConfig, compose_folder, UNKNOWN_DATE_FOLDER
from rename_engine import (
    CONCERN_ERROR,
    CONCERN_INFO,
    CONCERN_WARN,
    CONCERN_FILL_RED,
    CONCERN_FILL_YELLOW,
    _append_concerns,
    _column_index_map,
    _coerce_datetime,
    _ensure_column,
    _row_dict,
    render_row,
)
from rollback import RollbackWriter


LogCallback      = Callable[[str], None]
# Progress callback shape matches CatalogWorker's Qt signal for the GUI:
# (current, total, message). message is typically the current filename
# being operated on so the progress widget can show it alongside the bar.
ProgressCallback = Callable[[int, int, str], None]


# Sidecar files that follow the primary to its destination. Lowercased,
# including the leading dot. Sourced from the design discussion — cover
# the common raw-pair / edit-metadata / thumbnail files seen in real
# mixed photo drives.
_SIDECAR_EXTENSIONS: Tuple[str, ...] = (
    ".xmp",    # Adobe / metadata
    ".aae",    # Apple Photos edit sidecar
    ".thm",    # Camera thumbnail
    ".dop",    # DxO PhotoLab
)

# Status values written into File_Status. Kept here so callers have a
# single source of truth.
STATUS_PENDING      = "Pending"
STATUS_COPIED       = "Copied"
STATUS_SKIPPED      = "Skipped"
STATUS_DUPEMOVED    = "DupeMoved"
STATUS_DUPEDELETED  = "DupeDeleted"
STATUS_ROLLBACK     = "Rollback"


# ---------------------------------------------------------------------------
# Cancellation plumbing
# ---------------------------------------------------------------------------
class OperationCancelled(Exception):
    """Raised internally when the caller sets the cancel_event."""


def _check_cancel(cancel_event: Optional[threading.Event]) -> None:
    if cancel_event is not None and cancel_event.is_set():
        raise OperationCancelled()


def _noop_log(_m: str) -> None:
    pass


def _noop_progress(_c: int, _t: int, _m: str = "") -> None:
    pass


# ---------------------------------------------------------------------------
# Sidecar discovery
# ---------------------------------------------------------------------------
def find_sidecars(primary_path: str) -> List[str]:
    """
    Return absolute paths of same-stem sidecars in the primary's
    folder. Sidecar extensions come from :data:`_SIDECAR_EXTENSIONS`.

    Matching is stem-exact (case-insensitive). We also include any
    same-stem JPG / JPEG that pairs with a RAW primary — the most
    common "RAW+JPEG" shoot mode where both files are the same
    photograph rather than edits of it.
    """
    try:
        parent = os.path.dirname(primary_path)
        stem = Path(primary_path).stem
        primary_ext = Path(primary_path).suffix.lower()
    except (TypeError, ValueError):
        return []

    if not parent or not os.path.isdir(parent):
        return []

    sidecars: List[str] = []
    try:
        entries = os.listdir(parent)
    except OSError:
        return []

    raw_extensions = {".cr2", ".cr3", ".nef", ".arw", ".dng", ".orf", ".rw2"}
    primary_is_raw = primary_ext in raw_extensions

    for name in entries:
        cand = os.path.join(parent, name)
        if cand == primary_path:
            continue
        if not os.path.isfile(cand):
            continue
        c_stem = Path(name).stem
        c_ext = Path(name).suffix.lower()
        if c_stem.lower() != stem.lower():
            continue
        if c_ext in _SIDECAR_EXTENSIONS:
            sidecars.append(cand)
        elif primary_is_raw and c_ext in (".jpg", ".jpeg"):
            sidecars.append(cand)
    return sidecars


# ---------------------------------------------------------------------------
# Destination path resolution
# ---------------------------------------------------------------------------
def _available_path(target: str) -> str:
    """
    Return *target* if it doesn't exist; otherwise append ``_2``,
    ``_3``, ... before the extension until an unused name is found.
    """
    if not os.path.exists(target):
        return target
    stem, ext = os.path.splitext(target)
    n = 2
    while True:
        candidate = f"{stem}_{n}{ext}"
        if not os.path.exists(candidate):
            return candidate
        n += 1


# ---------------------------------------------------------------------------
# Populate destination columns (called during the catalog pipeline)
# ---------------------------------------------------------------------------
def populate_destination_columns(
    rows: List[Dict],
    destination_root: str,
    folder_config: FolderConfig,
    rename_template: str,
) -> Dict:
    """
    For every row in *rows*, compute and store:

        File_DestFolder  — <destination_root>\<composed relative folder>
        File_DestPath    — File_DestFolder \ File_RenameName
        File_RenameName  — rendered from *rename_template* (if the
                           template is populated and renders ok; rows
                           that fail fall back to the original
                           File_Name so the copy still has a target).

    Mutates rows in place. Returns a summary dict with counts of
    rows successfully resolved, rows that fell back to File_Name,
    and rows flagged with Unknown_Date folder.
    """
    summary = {
        "resolved": 0,
        "rename_fallback": 0,
        "unknown_date": 0,
        "dupe_fallback_date": 0,
    }

    dest_root = os.path.abspath(destination_root) if destination_root else ""

    for row in rows:
        # ---- 1. Rendered filename (rename) -----------------------------
        rendered_name = None
        if rename_template and rename_template.strip():
            try:
                rendered_name, reason, used_fallback = render_row(
                    rename_template, row,
                )
                if used_fallback:
                    summary["dupe_fallback_date"] += 1
                    concerns = row.setdefault("_concerns", [])
                    msg = f"{CONCERN_INFO} Used File_Date (EXIF DateTimeOriginal missing)"
                    if msg not in concerns:
                        concerns.append(msg)
                if rendered_name is None and reason:
                    # Template couldn't render — fall back to original name
                    # so the copy pass still has something to write.
                    summary["rename_fallback"] += 1
                    row.setdefault("_concerns", []).append(
                        f"{CONCERN_WARN} Rename fell back to original filename ({reason})"
                    )
            except Exception as e:  # noqa: BLE001
                logging.warning("render_row raised for %s: %s", row.get("File_Path"), e)
                rendered_name = None

        if not rendered_name:
            rendered_name = row.get("File_Name")
        row["File_RenameName"] = rendered_name

        # ---- 2. Destination folder (from checkboxes) -------------------
        # Prefer DateTimeOriginal, fall back to File_Date (same logic
        # the rename engine uses for %Date_*% tokens).
        dt = _coerce_datetime(row.get("DateTimeOriginal"))
        if dt is None:
            dt_fd = row.get("File_Date")
            if isinstance(dt_fd, datetime):
                dt = dt_fd

        rel_folder = compose_folder(folder_config, dt)
        if rel_folder == UNKNOWN_DATE_FOLDER:
            summary["unknown_date"] += 1
            row.setdefault("_concerns", []).append(
                f"{CONCERN_WARN} No usable date for folder composition — placed in Unknown_Date"
            )

        if dest_root:
            if rel_folder:
                dest_folder = os.path.join(dest_root, rel_folder)
            else:
                dest_folder = dest_root
            row["File_DestFolder"] = dest_folder
            row["File_DestPath"] = os.path.join(dest_folder, rendered_name) if rendered_name else ""
            if row["File_DestPath"]:
                summary["resolved"] += 1
        else:
            # No destination chosen — leave File_Dest* blank but still
            # expose the rendered folder so the user sees what *would*
            # have been produced.
            row["File_DestFolder"] = rel_folder
            row["File_DestPath"] = ""

        # File_Status defaults to Pending once a destination is resolved.
        # Use a plain assign rather than setdefault because extract_metadata
        # pre-populates File_Status=None, which would block setdefault from
        # overwriting it.
        if row.get("File_DestPath"):
            if not row.get("File_Status"):
                row["File_Status"] = STATUS_PENDING

    return summary


# ---------------------------------------------------------------------------
# Copy pass
# ---------------------------------------------------------------------------
def copy_to_destination(
    xlsx_path: str,
    destination_folder: str,
    progress_callback: Optional[ProgressCallback] = None,
    log_callback: Optional[LogCallback] = None,
    cancel_event: Optional[threading.Event] = None,
) -> Dict:
    """
    Copy every row's source file (``File_Path``) to its computed
    ``File_DestPath``, following same-stem sidecars to the same
    destination folder. Writes a ``_rollback_*.jsonl`` in
    *destination_folder* and updates ``File_Status`` / ``File_DestPath``
    (with collision suffixes if applicable) back into the workbook.

    Returns a summary dict: ``copied``, ``skipped``, ``sidecars_copied``,
    ``errors``, ``cancelled``, ``journal_path``.
    """
    log  = log_callback or _noop_log
    prog = progress_callback or _noop_progress

    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(xlsx_path)
    if not destination_folder:
        raise ValueError("destination_folder is required")

    os.makedirs(destination_folder, exist_ok=True)

    log(f"Opening workbook: {xlsx_path}")
    wb = load_workbook(xlsx_path)
    summary = {
        "copied": 0,
        "skipped": 0,
        "sidecars_copied": 0,
        "errors": 0,
        "cancelled": False,
        "journal_path": None,
    }

    try:
        ws = wb.active
        col_map = _column_index_map(ws)
        path_col    = _ensure_column(ws, col_map, "File_Path")
        dest_col    = _ensure_column(ws, col_map, "File_DestPath")
        dfolder_col = _ensure_column(ws, col_map, "File_DestFolder")
        status_col  = _ensure_column(ws, col_map, "File_Status")
        concern_col = _ensure_column(ws, col_map, "File_Concern")

        total = max(0, ws.max_row - 1)
        with RollbackWriter(destination_folder, tag="copy") as rb:
            summary["journal_path"] = str(rb.path)

            try:
                for i, row_idx in enumerate(range(2, ws.max_row + 1), start=1):
                    _check_cancel(cancel_event)
                    src = ws.cell(row=row_idx, column=path_col).value
                    dst = ws.cell(row=row_idx, column=dest_col).value
                    existing_status = ws.cell(row=row_idx, column=status_col).value

                    if existing_status == STATUS_COPIED:
                        summary["skipped"] += 1
                        continue
                    if not src or not dst:
                        summary["skipped"] += 1
                        _append_concerns(ws, row_idx, concern_col, [
                            f"{CONCERN_WARN} Copy skipped: File_Path or File_DestPath blank"
                        ])
                        ws.cell(row=row_idx, column=status_col, value=STATUS_SKIPPED)
                        continue
                    if not os.path.exists(src):
                        summary["errors"] += 1
                        _append_concerns(ws, row_idx, concern_col, [
                            f"{CONCERN_ERROR} Copy failed: source no longer exists"
                        ])
                        ws.cell(row=row_idx, column=status_col, value=STATUS_SKIPPED)
                        continue

                    # Ensure the destination folder exists.
                    dest_dir = os.path.dirname(dst)
                    pre_existing_dir = os.path.isdir(dest_dir)
                    try:
                        os.makedirs(dest_dir, exist_ok=True)
                        if not pre_existing_dir:
                            rb.record("mkdir", src="", dst=dest_dir)
                    except OSError as e:
                        summary["errors"] += 1
                        _append_concerns(ws, row_idx, concern_col, [
                            f"{CONCERN_ERROR} Could not create destination folder: {e}"
                        ])
                        ws.cell(row=row_idx, column=status_col, value=STATUS_SKIPPED)
                        continue

                    # Resolve collision on the destination file itself.
                    final_dst = _available_path(dst)
                    if final_dst != dst:
                        _append_concerns(ws, row_idx, concern_col, [
                            f"{CONCERN_WARN} Destination existed; copied as {os.path.basename(final_dst)}"
                        ])
                        # Update File_DestPath to reflect what actually happened.
                        ws.cell(row=row_idx, column=dest_col, value=final_dst)

                    # Copy primary.
                    try:
                        shutil.copy2(src, final_dst)
                        rb.record("copy", src=src, dst=final_dst)
                        summary["copied"] += 1
                    except OSError as e:
                        summary["errors"] += 1
                        _append_concerns(ws, row_idx, concern_col, [
                            f"{CONCERN_ERROR} Copy failed: {e}"
                        ])
                        ws.cell(row=row_idx, column=status_col, value=STATUS_SKIPPED)
                        continue

                    # Sidecars follow the primary.
                    sidecars = find_sidecars(src)
                    for sc in sidecars:
                        sc_name = Path(sc).name
                        # Use the renamed stem from the primary dest so a
                        # "IMG_0001.NEF → 2019-06-15_IMG_0001.NEF" rename
                        # keeps the sidecar's stem aligned.
                        primary_stem = Path(final_dst).stem
                        sc_ext = Path(sc_name).suffix
                        sc_dst = os.path.join(dest_dir, primary_stem + sc_ext)
                        sc_dst = _available_path(sc_dst)
                        try:
                            shutil.copy2(sc, sc_dst)
                            rb.record("copy", src=sc, dst=sc_dst, sidecar_of=src)
                            summary["sidecars_copied"] += 1
                        except OSError as e:
                            _append_concerns(ws, row_idx, concern_col, [
                                f"{CONCERN_WARN} Sidecar copy failed ({sc_name}): {e}"
                            ])

                    ws.cell(row=row_idx, column=status_col, value=STATUS_COPIED)

                    if i % 25 == 0 or i == total:
                        prog(i, total, os.path.basename(src))
                        log(f"  copied {i}/{total}: {os.path.basename(src)}")
            except OperationCancelled:
                summary["cancelled"] = True
                log("Copy cancelled by user.")

        log(
            f"Copy complete: {summary['copied']:,} primaries, "
            f"{summary['sidecars_copied']:,} sidecars, "
            f"{summary['skipped']:,} skipped, {summary['errors']:,} error(s)."
        )
        log(f"Rollback journal: {summary['journal_path']}")
        wb.save(xlsx_path)
    finally:
        wb.close()

    return summary


# ---------------------------------------------------------------------------
# Move / Delete non-keepers
# ---------------------------------------------------------------------------
def _iter_rows_matching_keep(ws, col_map: Dict[str, int], keep_value: bool):
    """
    Yield (row_idx, row_dict) pairs whose File_DupeKeep == keep_value.

    Rows with File_DupeKeep blank/None are **never** yielded — those
    aren't part of any duplicate group and must not be swept up by the
    non-keeper move/delete actions. That means this function returns a
    strict subset of "rows where the user has a clear keep/non-keep
    answer", which is exactly what the destructive buttons need.
    """
    for row_idx in range(2, ws.max_row + 1):
        row = _row_dict(ws, row_idx, col_map)
        kv = row.get("File_DupeKeep")

        # Explicit None / blank → skip entirely. A row that never went
        # through dupe detection has no business being moved or deleted.
        if kv is None:
            continue
        if isinstance(kv, str):
            kv_stripped = kv.strip().lower()
            if kv_stripped in ("", "none"):
                continue
            kv_bool = kv_stripped in ("true", "1", "yes", "y")
        else:
            kv_bool = bool(kv)
        if bool(kv_bool) is bool(keep_value):
            yield row_idx, row


def move_non_keepers(
    xlsx_path: str,
    holding_folder: str,
    progress_callback: Optional[ProgressCallback] = None,
    log_callback: Optional[LogCallback] = None,
    cancel_event: Optional[threading.Event] = None,
) -> Dict:
    """
    Move every row with ``File_DupeKeep = FALSE`` from its source
    location into *holding_folder* so the user can review duplicates
    in one place before deciding what to delete.

    Relative paths under the source root are preserved so the
    holding folder ends up looking like a mirror of the source
    for the subset of non-keeper files only.
    """
    log  = log_callback or _noop_log
    prog = progress_callback or _noop_progress

    os.makedirs(holding_folder, exist_ok=True)

    wb = load_workbook(xlsx_path)
    summary = {"moved": 0, "skipped": 0, "errors": 0, "cancelled": False,
               "journal_path": None}
    try:
        ws = wb.active
        col_map = _column_index_map(ws)
        status_col  = _ensure_column(ws, col_map, "File_Status")
        concern_col = _ensure_column(ws, col_map, "File_Concern")

        targets = list(_iter_rows_matching_keep(ws, col_map, keep_value=False))
        total = len(targets)
        log(f"Move non-keepers: {total} row(s) flagged File_DupeKeep=FALSE")

        with RollbackWriter(holding_folder, tag="move") as rb:
            summary["journal_path"] = str(rb.path)
            try:
                for i, (row_idx, row) in enumerate(targets, start=1):
                    _check_cancel(cancel_event)
                    src = row.get("File_Path")
                    if not src or not os.path.exists(src):
                        summary["skipped"] += 1
                        continue

                    # Keep the non-keeper's filename + a shallow relative
                    # structure so two files of the same name from
                    # different source folders don't collide in holding.
                    rel = os.path.basename(os.path.dirname(src))
                    dst_dir = os.path.join(holding_folder, rel) if rel else holding_folder
                    os.makedirs(dst_dir, exist_ok=True)
                    dst = _available_path(os.path.join(dst_dir, os.path.basename(src)))
                    try:
                        shutil.move(src, dst)
                        rb.record("move", src=src, dst=dst)
                        ws.cell(row=row_idx, column=status_col, value=STATUS_DUPEMOVED)
                        summary["moved"] += 1
                    except OSError as e:
                        summary["errors"] += 1
                        _append_concerns(ws, row_idx, concern_col, [
                            f"{CONCERN_ERROR} Move non-keeper failed: {e}"
                        ])
                    if i % 25 == 0 or i == total:
                        prog(i, total, os.path.basename(src))
            except OperationCancelled:
                summary["cancelled"] = True
                log("Move non-keepers cancelled by user.")

        log(
            f"Move non-keepers complete: {summary['moved']:,} moved, "
            f"{summary['skipped']:,} skipped, {summary['errors']:,} error(s)."
        )
        log(f"Rollback journal: {summary['journal_path']}")
        wb.save(xlsx_path)
    finally:
        wb.close()
    return summary


def delete_non_keepers(
    xlsx_path: str,
    rollback_dir: str,
    progress_callback: Optional[ProgressCallback] = None,
    log_callback: Optional[LogCallback] = None,
    cancel_event: Optional[threading.Event] = None,
) -> Dict:
    """
    Delete every row with ``File_DupeKeep = FALSE`` from the source
    drive. Writes a rollback journal in *rollback_dir* (usually the
    destination folder) noting each deletion — note that delete is
    not truly auto-undoable (the bytes are gone), but the journal
    preserves the paths so the user has a paper trail for recovery
    from backup.
    """
    log  = log_callback or _noop_log
    prog = progress_callback or _noop_progress

    os.makedirs(rollback_dir, exist_ok=True)

    wb = load_workbook(xlsx_path)
    summary = {"deleted": 0, "skipped": 0, "errors": 0, "cancelled": False,
               "journal_path": None}
    try:
        ws = wb.active
        col_map = _column_index_map(ws)
        status_col  = _ensure_column(ws, col_map, "File_Status")
        concern_col = _ensure_column(ws, col_map, "File_Concern")

        targets = list(_iter_rows_matching_keep(ws, col_map, keep_value=False))
        total = len(targets)
        log(f"Delete non-keepers: {total} row(s) flagged File_DupeKeep=FALSE")

        with RollbackWriter(rollback_dir, tag="delete") as rb:
            summary["journal_path"] = str(rb.path)
            try:
                for i, (row_idx, row) in enumerate(targets, start=1):
                    _check_cancel(cancel_event)
                    src = row.get("File_Path")
                    if not src or not os.path.exists(src):
                        summary["skipped"] += 1
                        continue
                    try:
                        os.remove(src)
                        rb.record("delete", src=src)
                        ws.cell(row=row_idx, column=status_col, value=STATUS_DUPEDELETED)
                        summary["deleted"] += 1
                    except OSError as e:
                        summary["errors"] += 1
                        _append_concerns(ws, row_idx, concern_col, [
                            f"{CONCERN_ERROR} Delete non-keeper failed: {e}"
                        ])
                    if i % 25 == 0 or i == total:
                        prog(i, total, os.path.basename(src))
            except OperationCancelled:
                summary["cancelled"] = True
                log("Delete non-keepers cancelled by user.")

        log(
            f"Delete non-keepers complete: {summary['deleted']:,} deleted, "
            f"{summary['skipped']:,} skipped, {summary['errors']:,} error(s)."
        )
        log(f"Rollback journal: {summary['journal_path']}")
        wb.save(xlsx_path)
    finally:
        wb.close()
    return summary


# ---------------------------------------------------------------------------
# Empty source folder cleanup
# ---------------------------------------------------------------------------
def find_empty_source_folders(source_folder: str) -> List[str]:
    """
    Walk *source_folder* bottom-up and return every directory that
    contains no files (hidden or otherwise) in its own subtree.

    Excludes *source_folder* itself from the candidate list — we
    don't want to let the user accidentally remove the root they
    scanned.
    """
    candidates: List[str] = []
    source_abs = os.path.abspath(source_folder)
    for root, dirs, files in os.walk(source_abs, topdown=False):
        if os.path.abspath(root) == source_abs:
            continue
        # A directory is empty if it has no files AND no
        # subdirectories (or only subdirectories we've already
        # added because they themselves were empty — bottom-up walk
        # means subdirs were already considered).
        try:
            remaining = os.listdir(root)
        except OSError:
            continue
        if not remaining:
            candidates.append(root)
    return candidates


def remove_empty_source_folders(
    folders: List[str],
    rollback_dir: str,
    log_callback: Optional[LogCallback] = None,
) -> Dict:
    """
    Remove every directory in *folders* (assumed empty). Writes a
    rollback journal entry for each removal so Undo can recreate
    them. Silently skips any directory that has become non-empty
    between find_empty_source_folders and now.
    """
    log = log_callback or _noop_log
    summary = {"removed": 0, "skipped": 0, "errors": 0, "journal_path": None}

    if not folders:
        log("No empty source folders to remove.")
        return summary

    os.makedirs(rollback_dir, exist_ok=True)
    with RollbackWriter(rollback_dir, tag="cleanup") as rb:
        summary["journal_path"] = str(rb.path)
        for folder in folders:
            try:
                if not os.path.isdir(folder):
                    summary["skipped"] += 1
                    continue
                if os.listdir(folder):
                    # Re-check — maybe a background process dropped a
                    # file there since find_empty_source_folders walked.
                    summary["skipped"] += 1
                    continue
                os.rmdir(folder)
                rb.record("rmdir", src=folder)
                summary["removed"] += 1
            except OSError as e:
                log(f"  could not remove {folder}: {e}")
                summary["errors"] += 1

    log(
        f"Source cleanup: {summary['removed']:,} folder(s) removed, "
        f"{summary['skipped']:,} skipped, {summary['errors']:,} error(s)."
    )
    return summary
